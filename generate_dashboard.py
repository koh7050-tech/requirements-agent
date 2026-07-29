#!/usr/bin/env python3
"""XLSX 요구분석 데이터팩 → HTML 대시보드 생성 (다크 모던 + 인포 팝업)"""

from __future__ import annotations
import sys
from pathlib import Path
from datetime import date

try:
    import openpyxl
except ImportError:
    print("[ERROR] pip3 install openpyxl")
    sys.exit(1)


def load_sheet(wb, name: str) -> tuple[list[str], list[list]]:
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h or "") for h in rows[0]]
    data = [[str(c or "") for c in r] for r in rows[1:] if any(c for c in r)]
    return headers, data


def badge(text: str) -> str:
    colors = {
        "Must":   ("#c0392b", "#fff"),
        "Should": ("#b07d10", "#fff"),
        "Could":  ("#1a5fa8", "#fff"),
        "Won't":  ("#444", "#aaa"),
        "PASS":   ("#1a6b3a", "#3ecf8e"),
        "FAIL":   ("#5c1a1a", "#f07070"),
        "대기":   ("#2a2a2a", "#888"),
        "진행중": ("#5c3d00", "#f39c12"),
        "완료":   ("#1a6b3a", "#3ecf8e"),
        "Active": ("#1a3a5c", "#5dade2"),
        "기능":   ("#3b1a5c", "#c39bd3"),
        "비기능": ("#0d3d33", "#48c9b0"),
    }
    bg, fg = colors.get(text, ("#2a2a2a", "#aaa"))
    return (f'<span style="background:{bg};color:{fg};padding:2px 9px;'
            f'border-radius:20px;font-size:11px;font-weight:600;'
            f'white-space:nowrap;border:1px solid {fg}22">{text}</span>')


def table_html(headers: list[str], rows: list[list], badge_cols: list[str] = None) -> str:
    badge_cols = badge_cols or []
    th = "".join(f"<th>{h}</th>" for h in headers)
    trs = []
    for i, row in enumerate(rows):
        tds = []
        for h, v in zip(headers, row):
            if h in badge_cols and v:
                tds.append(f"<td>{badge(v)}</td>")
            else:
                tds.append(f"<td>{v}</td>")
        trs.append(f'<tr class="{"row-alt" if i % 2 else ""}">' + "".join(tds) + "</tr>")
    return f"""
<div class="table-wrap">
<table>
  <thead><tr>{th}</tr></thead>
  <tbody>{"".join(trs)}</tbody>
</table>
</div>"""


def kpi_cards(headers: list[str], rows: list[list]) -> str:
    if not rows:
        return '<p class="empty">KPI 데이터 없음</p>'
    idx = {h: i for i, h in enumerate(headers)}
    cards = []
    for r in rows:
        kpi_id = r[idx.get("KPI_ID", 0)]
        req    = r[idx.get("연결_REQ", 1)]
        name   = r[idx.get("지표명", 3)]
        target = r[idx.get("목표값", 8)]
        unit   = r[idx.get("단위", 4)]
        period = r[idx.get("측정주기", 6)]
        owner  = r[idx.get("오너", 10)]
        cards.append(f"""
<div class="kpi-card">
  <div class="kpi-id">{kpi_id} · {req}</div>
  <div class="kpi-name">{name}</div>
  <div class="kpi-target">{target} <span class="kpi-unit">{unit}</span></div>
  <div class="kpi-meta">{period} 측정 · {owner}</div>
</div>""")
    return '<div class="kpi-grid">' + "".join(cards) + "</div>"


def moscow_chart(headers: list[str], rows: list[list]) -> str:
    if not rows:
        return ""
    idx = {h: i for i, h in enumerate(headers)}
    counts: dict[str, int] = {"Must": 0, "Should": 0, "Could": 0, "Won't": 0}
    for r in rows:
        c = r[idx.get("분류", 1)]
        if c in counts:
            counts[c] += 1
    total = sum(counts.values()) or 1
    colors = {"Must": "#c0392b", "Should": "#d4a017", "Could": "#2980b9", "Won't": "#555"}
    bars = ""
    for label, count in counts.items():
        pct = count / total * 100
        bars += f"""
<div class="bar-row">
  <div class="bar-label">{label}</div>
  <div class="bar-track">
    <div class="bar-fill" style="width:{pct:.1f}%;background:{colors[label]}"></div>
  </div>
  <div class="bar-count">{count}</div>
</div>"""
    return f'<div class="bar-chart">{bars}</div>'


def action_timeline(headers: list[str], rows: list[list]) -> str:
    if not rows:
        return '<p class="empty">액션 없음</p>'
    idx = {h: i for i, h in enumerate(headers)}
    by_month: dict[str, list] = {}
    for r in rows:
        due = r[idx.get("목표일", 4)]
        month = due[:7] if len(due) >= 7 else "미정"
        by_month.setdefault(month, []).append(r)
    html = ""
    for month in sorted(by_month):
        items = ""
        for r in by_month[month]:
            act_id = r[idx.get("ACT_ID", 0)]
            action = r[idx.get("액션", 2)]
            owner  = r[idx.get("오너", 3)]
            status = r[idx.get("상태", 5)]
            pri    = r[idx.get("우선순위", 6)]
            items += f"""
<div class="tl-item">
  <span class="tl-id">{act_id}</span>
  {badge(pri)} {badge(status)}
  <span class="tl-action">{action}</span>
  <span class="tl-owner">· {owner}</span>
</div>"""
        html += f'<div class="tl-month"><div class="tl-month-label">{month}</div>{items}</div>'
    return f'<div class="timeline">{html}</div>'


INFO = {
    "fr": {
        "title": "기능 요구사항 (FR)",
        "rows": [
            ("정의", "시스템이 반드시 수행해야 하는 구체적인 기능을 정의한 요구사항입니다."),
            ("포함 항목", "ID · 요구사항 설명 · 이해관계자 · 우선순위(MoSCoW) · 수용기준 · 원시ID · 출처ID · 상태"),
            ("활용 방법", "우선순위 컬럼으로 정렬하여 Must → Should → Could 순서로 개발 스코프를 결정합니다."),
            ("주의사항", "Must 항목은 반드시 1차 릴리즈에 포함되어야 합니다."),
        ]
    },
    "nfr": {
        "title": "비기능 요구사항 (NFR)",
        "rows": [
            ("정의", "성능·보안·가용성·확장성 등 시스템의 품질 속성을 정의한 요구사항입니다."),
            ("포함 항목", "ID · 유형 · 요구사항 설명 · 이해관계자 · 우선순위 · 수용기준 · 상태"),
            ("활용 방법", "시스템 아키텍처 설계 시 비기능 요구사항을 기반으로 기술 스택과 인프라를 결정합니다."),
            ("주의사항", "성능 목표값(응답시간·처리량)은 구체적인 수치로 명시해야 합니다."),
        ]
    },
    "moscow": {
        "title": "MoSCoW 우선순위",
        "rows": [
            ("정의", "Must·Should·Could·Won't 4단계로 요구사항 우선순위를 분류하는 기법입니다."),
            ("Must", "이번 릴리즈에 반드시 포함되어야 하는 핵심 기능"),
            ("Should", "중요하지만 없어도 릴리즈 가능한 기능"),
            ("Could", "여유가 있을 때 구현하면 좋은 기능"),
            ("Won't", "이번 스코프에서는 제외하는 기능"),
        ]
    },
    "kpi": {
        "title": "KPI (핵심 성과 지표)",
        "rows": [
            ("정의", "프로젝트 성공 여부를 측정하는 정량적 지표입니다."),
            ("포함 항목", "KPI ID · 연결 요구사항 · 지표명 · 단위 · 목표값 · 측정주기 · 오너"),
            ("활용 방법", "프로젝트 완료 후 KPI 달성 여부를 주기적으로 모니터링합니다."),
            ("계산 기준", "목표값 대비 실제값의 달성률(%)로 산정합니다."),
        ]
    },
    "action": {
        "title": "액션 아이템",
        "rows": [
            ("정의", "프로젝트 진행을 위해 수행해야 할 구체적인 작업 목록입니다."),
            ("포함 항목", "ACT ID · 연결 요구사항 · 액션 내용 · 오너 · 목표일 · 상태 · 우선순위"),
            ("활용 방법", "목표일 기준으로 정렬하여 타임라인을 관리합니다."),
            ("상태 관리", "대기 → 진행중 → 완료 순서로 상태를 업데이트합니다."),
        ]
    },
    "trace": {
        "title": "추적성 매트릭스",
        "rows": [
            ("정의", "인터뷰 → 원시요구사항 → FR/NFR 간의 연결 관계를 추적하는 매트릭스입니다."),
            ("목적", "요구사항의 출처를 역추적하고, 누락된 요구사항을 식별합니다."),
            ("활용 방법", "변경 요청 발생 시 영향받는 요구사항을 추적하여 변경 범위를 파악합니다."),
        ]
    },
    "qc": {
        "title": "품질 Gate",
        "rows": [
            ("정의", "요구사항 분석의 완성도와 품질을 검증하는 체크리스트입니다."),
            ("PASS 기준", "모든 필수 항목이 작성되고 일관성이 확인된 경우"),
            ("FAIL 처리", "FAIL 항목은 반드시 수정 후 재검토해야 합니다."),
            ("필요 이유", "불완전한 요구사항이 개발 단계로 넘어가는 것을 방지합니다."),
        ]
    },
}


def info_btn(key: str) -> str:
    info = INFO.get(key, {})
    title = info.get("title", "")
    rows_html = ""
    for k, v in info.get("rows", []):
        rows_html += f'<div class="ip-row"><div class="ip-key">{k}</div><div class="ip-val">{v}</div></div>'
    return f"""
<div class="info-wrap">
  <button class="info-btn" onclick="toggleInfo('{key}', this)">i</button>
  <div class="info-popup" id="popup-{key}">
    <h4>{title}</h4>
    {rows_html}
  </div>
</div>"""


def section(title: str, content: str, info_key: str = "") -> str:
    ib = info_btn(info_key) if info_key else ""
    return f"""
<section class="section">
  <div class="section-header">
    <h2 class="section-title">{title}</h2>
    {ib}
  </div>
  {content}
</section>"""


def generate(xlsx_path: Path, output_path: Path, project: str, mode: str = "") -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    today = date.today().isoformat()

    _, src_rows     = load_sheet(wb, "1_출처")
    fr_h, fr_rows   = load_sheet(wb, "3_기능요구사항")
    nfr_h, nfr_rows = load_sheet(wb, "4_비기능요구사항")
    mo_h, mo_rows   = load_sheet(wb, "5_분석_MoSCoW")
    kpi_h, kpi_rows = load_sheet(wb, "6_KPI")
    act_h, act_rows = load_sheet(wb, "7_액션아이템")
    tr_h, tr_rows   = load_sheet(wb, "8_추적성")
    qc_h, qc_rows   = load_sheet(wb, "9_품질Gate")

    qc_pass  = sum(1 for r in qc_rows if len(r) > 1 and r[1] == "PASS")
    qc_total = len(qc_rows)

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Automated Requirements Analysis — {project}</title>
<style>
:root{{
  --bg:#0d0d0d;--surface:#161616;--surface2:#1e1e1e;
  --border:#2a2a2a;--border2:#383838;
  --text:#f0f0f0;--muted:#777;--muted2:#444;
  --green:#3ecf8e;--red:#f04040;--radius:10px;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Apple SD Gothic Neo',sans-serif;
  background:var(--bg);color:var(--text);font-size:13px;line-height:1.6}}
.topbar{{display:flex;justify-content:space-between;align-items:center;
  padding:0 32px;height:54px;border-bottom:1px solid var(--border);
  background:rgba(13,13,13,.96);position:sticky;top:0;z-index:100;backdrop-filter:blur(10px)}}
.topbar-label{{font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--muted)}}
.topbar-meta{{font-size:11px;color:var(--muted)}}
.hero{{padding:40px 32px 24px}}
.hero h1{{font-size:2rem;font-weight:700;letter-spacing:-.4px;margin-bottom:6px}}
.hero-sub{{font-size:13px;color:var(--muted)}}
.mode-tag{{display:inline-block;margin-top:10px;font-size:11px;padding:3px 12px;
  border:1px solid var(--border2);border-radius:20px;color:var(--muted)}}
.summary{{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;padding:0 32px 28px}}
.stat-card{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:18px 16px}}
.stat-num{{font-size:28px;font-weight:800;color:var(--text);letter-spacing:-.5px}}
.stat-num.green{{color:var(--green)}}
.stat-label{{font-size:11px;color:var(--muted);margin-top:4px;font-weight:500}}
.main{{padding:0 32px 48px}}
.section{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:22px 24px;margin-bottom:14px}}
.section-header{{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;padding-bottom:10px;border-bottom:1px solid var(--border)}}
.section-title{{font-size:13px;font-weight:700;color:var(--text);letter-spacing:.1px}}
.table-wrap{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{background:var(--surface2);color:var(--muted);padding:8px 10px;text-align:left;
  font-weight:600;white-space:nowrap;font-size:11px;letter-spacing:.4px;text-transform:uppercase;
  border-bottom:1px solid var(--border2)}}
td{{padding:8px 10px;border-bottom:1px solid var(--border);vertical-align:top;line-height:1.5;color:#ccc}}
.row-alt td{{background:rgba(255,255,255,.015)}}
tr:hover td{{background:rgba(255,255,255,.03)}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:12px}}
.kpi-card{{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:14px}}
.kpi-id{{font-size:10px;color:var(--muted);margin-bottom:4px;font-family:monospace}}
.kpi-name{{font-size:12px;font-weight:600;color:var(--text);margin-bottom:10px;line-height:1.4}}
.kpi-target{{font-size:24px;font-weight:800;color:var(--green)}}
.kpi-unit{{font-size:13px;font-weight:400;color:var(--muted)}}
.kpi-meta{{font-size:10px;color:var(--muted2);margin-top:6px}}
.bar-chart{{padding:4px 0}}
.bar-row{{display:flex;align-items:center;margin-bottom:10px;gap:10px}}
.bar-label{{width:56px;font-size:12px;font-weight:600;color:var(--text)}}
.bar-track{{flex:1;background:var(--surface2);border-radius:3px;height:16px;overflow:hidden;border:1px solid var(--border)}}
.bar-fill{{height:100%;border-radius:3px}}
.bar-count{{width:28px;font-size:12px;color:var(--muted);text-align:right}}
.timeline{{}}
.tl-month{{margin-bottom:16px}}
.tl-month-label{{font-size:11px;font-weight:700;color:var(--muted);background:var(--surface2);
  border:1px solid var(--border);padding:3px 10px;border-radius:20px;margin-bottom:10px;
  display:inline-block;font-family:monospace}}
.tl-item{{padding:7px 0 7px 14px;border-left:1px solid var(--border2);margin-bottom:4px;
  display:flex;align-items:center;flex-wrap:wrap;gap:6px}}
.tl-id{{font-size:10px;color:var(--muted);width:52px;font-family:monospace}}
.tl-action{{font-size:12px;color:#ccc}}
.tl-owner{{font-size:11px;color:var(--muted)}}
.two-col{{display:grid;grid-template-columns:1fr 1fr;gap:20px}}
.empty{{color:var(--muted);font-size:12px;padding:8px 0}}
footer{{text-align:center;padding:24px;font-size:11px;color:var(--muted2);border-top:1px solid var(--border)}}

/* 인포 팝업 */
.info-wrap{{position:relative;display:inline-flex}}
.info-btn{{width:22px;height:22px;border-radius:50%;border:1px solid var(--border2);
  background:var(--surface2);color:var(--muted);font-size:11px;font-weight:700;
  cursor:pointer;display:flex;align-items:center;justify-content:center;
  transition:border-color .15s,color .15s;flex-shrink:0;user-select:none}}
.info-btn:hover{{border-color:#555;color:var(--text)}}
.info-btn.active{{border-color:var(--green);color:var(--green)}}
.info-popup{{display:none;position:absolute;right:0;top:28px;width:300px;
  background:#1a1a1a;border:1px solid var(--border2);border-radius:var(--radius);
  padding:16px;z-index:999;box-shadow:0 12px 32px rgba(0,0,0,.6)}}
.info-popup.open{{display:block}}
.info-popup h4{{font-size:13px;font-weight:700;color:var(--text);margin-bottom:12px;
  padding-bottom:10px;border-bottom:1px solid var(--border)}}
.ip-row{{margin-bottom:10px}}
.ip-row:last-child{{margin-bottom:0}}
.ip-key{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;
  color:#7b8cde;margin-bottom:3px}}
.ip-val{{font-size:12px;color:#bbb;line-height:1.55}}

@media(max-width:700px){{
  .summary{{grid-template-columns:repeat(2,1fr)}}
  .two-col{{grid-template-columns:1fr}}
  .topbar,.hero,.main{{padding-left:16px;padding-right:16px}}
  .summary{{padding-left:16px;padding-right:16px}}
}}
</style>
</head>
<body>

<div class="topbar">
  <span class="topbar-label">Analysis Result</span>
  <span class="topbar-meta">생성일: {today} · Requirements Analysis Agent{f" · {mode}" if mode else ""}</span>
</div>

<div class="hero">
  <h1>Automated Requirements Analysis</h1>
  <div class="hero-sub">{project}</div>
  {f'<span class="mode-tag">{mode}</span>' if mode else ""}
</div>

<div class="summary">
  <div class="stat-card"><div class="stat-num">{len(fr_rows)}</div><div class="stat-label">기능 요구사항 (FR)</div></div>
  <div class="stat-card"><div class="stat-num">{len(nfr_rows)}</div><div class="stat-label">비기능 요구사항 (NFR)</div></div>
  <div class="stat-card"><div class="stat-num">{len(kpi_rows)}</div><div class="stat-label">KPI</div></div>
  <div class="stat-card"><div class="stat-num">{len(act_rows)}</div><div class="stat-label">액션 아이템</div></div>
  <div class="stat-card"><div class="stat-num green">{qc_pass}/{qc_total}</div><div class="stat-label">품질 Gate PASS</div></div>
</div>

<div class="main">

{section("기능 요구사항 (FR)", table_html(fr_h, fr_rows, badge_cols=["우선순위","상태"]), "fr")}
{section("비기능 요구사항 (NFR)", table_html(nfr_h, nfr_rows, badge_cols=["우선순위","상태"]), "nfr")}

<section class="section">
  <div class="section-header">
    <h2 class="section-title">MoSCoW 우선순위 분석</h2>
    {info_btn("moscow")}
  </div>
  <div class="two-col">
    <div>{moscow_chart(mo_h, mo_rows)}</div>
    <div>{table_html(mo_h, mo_rows, badge_cols=["분류"])}</div>
  </div>
</section>

{section("KPI", kpi_cards(kpi_h, kpi_rows), "kpi")}
{section("액션 아이템 타임라인", action_timeline(act_h, act_rows), "action")}
{section("추적성 매트릭스", table_html(tr_h, tr_rows, badge_cols=["상태"]), "trace")}
{section("품질 Gate", table_html(qc_h, qc_rows, badge_cols=["결과"]), "qc")}

</div>

<footer>Requirements Analysis Agent · {today} · {project}</footer>

<script>
function toggleInfo(key, btn) {{
  var popup = document.getElementById('popup-' + key);
  var isOpen = popup.classList.contains('open');
  document.querySelectorAll('.info-popup').forEach(function(p) {{ p.classList.remove('open'); }});
  document.querySelectorAll('.info-btn').forEach(function(b) {{ b.classList.remove('active'); }});
  if (!isOpen) {{ popup.classList.add('open'); btn.classList.add('active'); }}
}}
document.addEventListener('click', function(e) {{
  if (!e.target.closest('.info-wrap')) {{
    document.querySelectorAll('.info-popup').forEach(function(p) {{ p.classList.remove('open'); }});
    document.querySelectorAll('.info-btn').forEach(function(b) {{ b.classList.remove('active'); }});
  }}
}});
</script>
</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")
    print(f"✓ 대시보드 생성: {output_path}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx",    default="outputs/요구분석_데이터팩.xlsx")
    parser.add_argument("--output",  default="outputs/dashboard.html")
    parser.add_argument("--project", default="요구사항분석")
    parser.add_argument("--mode",    default="")
    args = parser.parse_args()
    generate(Path(args.xlsx), Path(args.output), args.project, args.mode)
