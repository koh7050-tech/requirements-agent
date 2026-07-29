#!/usr/bin/env python3
"""XLSX → PRD 템플릿 HTML (인쇄 → PDF 변환용)"""

from __future__ import annotations
from datetime import date
from pathlib import Path


def load_sheet(wb, name):
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h or "") for h in rows[0]]
    data = [[str(c or "") for c in r] for r in rows[1:] if any(c for c in r)]
    return headers, data


def pri_badge(p: str) -> str:
    colors = {"Must": "#c0392b", "Should": "#e67e22", "Could": "#2980b9", "Won't": "#888"}
    c = colors.get(p, "#888")
    return f'<span style="color:{c};font-weight:700;font-size:11px">{p}</span>'


def generate_prd(xlsx_path: Path, project: str, mode: str = "") -> str:
    import openpyxl as ox
    wb = ox.load_workbook(xlsx_path)
    today = date.today().isoformat()

    fr_h,  fr_rows  = load_sheet(wb, "3_기능요구사항")
    nfr_h, nfr_rows = load_sheet(wb, "4_비기능요구사항")
    kpi_h, kpi_rows = load_sheet(wb, "6_KPI")
    act_h, act_rows = load_sheet(wb, "7_액션아이템")
    qc_h,  qc_rows  = load_sheet(wb, "9_품질Gate")

    qc_pass  = sum(1 for r in qc_rows if len(r) > 1 and r[1] == "PASS")
    qc_total = len(qc_rows)

    # ── FR 요구사항 행 ──────────────────────────────────
    fr_idx = {h: i for i, h in enumerate(fr_h)}
    fr_rows_html = ""
    for i, r in enumerate(fr_rows, 1):
        req = r[fr_idx.get("요구사항", 1)] if len(r) > 1 else ""
        ac  = r[fr_idx.get("수용기준", 4)] if len(r) > 4 else ""
        pri = r[fr_idx.get("우선순위", 3)] if len(r) > 3 else ""
        fr_rows_html += f"""<tr>
          <td class="num">{i}</td>
          <td>{req}</td>
          <td style="color:#555;font-size:12px">{ac}</td>
          <td>{pri_badge(pri)}</td>
          <td></td><td></td>
        </tr>"""

    # ── NFR 요구사항 행 ─────────────────────────────────
    nfr_idx = {h: i for i, h in enumerate(nfr_h)}
    nfr_rows_html = ""
    for i, r in enumerate(nfr_rows, 1):
        req = r[nfr_idx.get("요구사항", 2)]  if len(r) > 2 else ""
        typ = r[nfr_idx.get("유형", 1)]       if len(r) > 1 else ""
        pri = r[nfr_idx.get("우선순위", 4)]   if len(r) > 4 else ""
        nfr_rows_html += f"""<tr>
          <td class="num">{i}</td>
          <td>{req} <span style="color:#aaa;font-size:11px">[{typ}]</span></td>
          <td></td>
          <td>{pri_badge(pri)}</td>
          <td></td><td></td>
        </tr>"""

    # ── KPI → Success Metrics ────────────────────────────
    kpi_idx = {h: i for i, h in enumerate(kpi_h)}
    kpi_rows_html = ""
    for r in kpi_rows:
        name   = r[kpi_idx.get("지표명", 3)]  if len(r) > 3 else ""
        target = r[kpi_idx.get("목표값", 8)]  if len(r) > 8 else ""
        unit   = r[kpi_idx.get("단위", 4)]    if len(r) > 4 else ""
        kpi_rows_html += f"<tr><td>{name}</td><td>{target} {unit}</td></tr>"
    if not kpi_rows_html:
        kpi_rows_html = "<tr><td style='color:#aaa'>e.g., 시스템 응답시간</td><td style='color:#aaa'>e.g., 1초 이하</td></tr>"

    # ── 액션 아이템 행 ──────────────────────────────────
    act_idx = {h: i for i, h in enumerate(act_h)}
    act_rows_html = ""
    for i, r in enumerate(act_rows, 1):
        action = r[act_idx.get("액션", 2)]   if len(r) > 2 else ""
        owner  = r[act_idx.get("오너", 3)]   if len(r) > 3 else ""
        due    = r[act_idx.get("목표일", 4)] if len(r) > 4 else ""
        status = r[act_idx.get("상태", 5)]   if len(r) > 5 else ""
        act_rows_html += f"<tr><td class='num'>{i}</td><td>{action}</td><td>{owner}</td><td>{due}</td><td>{status}</td></tr>"
    if not act_rows_html:
        act_rows_html = "<tr><td class='num'>1</td><td style='color:#aaa'>e.g., 킥오프 미팅</td><td></td><td></td><td></td></tr>"

    # ── 품질 Gate 행 ────────────────────────────────────
    qc_rows_html = ""
    for r in qc_rows:
        gate   = r[0] if len(r) > 0 else ""
        result = r[1] if len(r) > 1 else ""
        desc   = r[2] if len(r) > 2 else ""
        color  = "#27ae60" if result == "PASS" else "#e74c3c"
        qc_rows_html += f'<tr><td>{gate}</td><td style="color:{color};font-weight:700">{result}</td><td>{desc}</td></tr>'

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PRD — {project}</title>
<style>
  @page {{ margin: 20mm 18mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: -apple-system, 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    background: #fff;
    color: #1a1a1a;
    font-size: 13px;
    line-height: 1.65;
  }}

  /* 인쇄 버튼 (화면에서만 표시) */
  .print-bar {{
    position: fixed; top: 0; left: 0; right: 0;
    background: #111; color: #fff;
    display: flex; align-items: center; justify-content: space-between;
    padding: 10px 24px; z-index: 999; font-size: 13px;
  }}
  .print-bar span {{ color: #aaa; font-size: 12px; }}
  .btn-print {{
    padding: 7px 20px; background: #3ecf8e; color: #000;
    border: none; border-radius: 6px; font-size: 13px;
    font-weight: 700; cursor: pointer;
  }}
  .btn-print:hover {{ opacity: .85; }}
  @media print {{ .print-bar {{ display: none !important; }} }}

  /* 문서 본문 */
  .doc {{
    max-width: 800px;
    margin: 0 auto;
    padding: 72px 32px 64px;
  }}
  @media print {{ .doc {{ padding: 0; }} }}

  /* 헤더 */
  .doc-title {{
    font-size: 1.9rem;
    font-weight: 800;
    color: #111;
    letter-spacing: -.4px;
    margin-bottom: 8px;
  }}
  .doc-desc {{
    font-size: 13px;
    color: #555;
    margin-bottom: 20px;
    line-height: 1.6;
  }}

  /* 섹션 제목 */
  h2 {{
    font-size: 1.15rem;
    font-weight: 700;
    color: #111;
    margin: 36px 0 6px;
    page-break-after: avoid;
  }}
  h2:first-of-type {{ margin-top: 24px; }}
  .section-desc {{
    font-size: 12px;
    color: #888;
    margin-bottom: 12px;
  }}

  /* Details 테이블 */
  .details-block {{
    border: 1px solid #e0e0e0;
    border-radius: 6px;
    overflow: hidden;
    margin: 10px 0 20px;
    font-size: 13px;
  }}
  .details-block table {{ width: 100%; border-collapse: collapse; }}
  .details-block td {{
    padding: 8px 14px;
    border-bottom: 1px solid #f0f0f0;
    vertical-align: middle;
  }}
  .details-block tr:last-child td {{ border-bottom: none; }}
  .details-block td:first-child {{
    color: #555; width: 180px; background: #fafafa; font-weight: 500;
  }}
  .badge-draft {{
    display: inline-block;
    font-size: 10px; font-weight: 700; letter-spacing: .5px;
    background: #f0f0f0; color: #555; padding: 2px 7px;
    border-radius: 3px; border: 1px solid #ddd;
  }}
  .badge-owner {{
    display: inline-block;
    font-size: 12px; font-weight: 600;
    background: #e8f0fe; color: #1967d2; padding: 2px 10px;
    border-radius: 12px;
  }}

  /* 일반 데이터 테이블 */
  .data-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 12px;
    margin: 10px 0 8px;
    border: 1px solid #e5e5e5;
    border-radius: 4px;
    overflow: hidden;
  }}
  .data-table th {{
    background: #f5f5f5;
    padding: 8px 12px;
    text-align: left;
    font-weight: 600;
    color: #444;
    font-size: 12px;
    border-bottom: 1px solid #e5e5e5;
    white-space: nowrap;
  }}
  .data-table td {{
    padding: 8px 12px;
    border-bottom: 1px solid #f0f0f0;
    vertical-align: top;
    color: #333;
    line-height: 1.5;
  }}
  .data-table tr:last-child td {{ border-bottom: none; }}
  .data-table tr:hover td {{ background: #fafafa; }}
  .data-table .num {{
    width: 28px; text-align: center;
    color: #aaa; font-size: 11px;
  }}

  /* 구분선 */
  .divider {{ border: none; border-top: 1px solid #f0f0f0; margin: 28px 0 0; }}

  /* 페이지 나누기 방지 */
  .section {{ page-break-inside: avoid; }}
</style>
</head>
<body>

<!-- 인쇄 버튼 바 (화면에서만 표시) -->
<div class="print-bar">
  <span>PDF로 저장하려면 → Ctrl+P(Windows) / Cmd+P(Mac) → "PDF로 저장" 선택</span>
  <button class="btn-print" onclick="window.print()">PDF로 저장</button>
</div>

<div class="doc">

  <!-- 타이틀 -->
  <div class="doc-title">Product requirements</div>
  <div class="doc-desc">
    Fill in project details in the table below. Generated by Requirements Analysis Agent.
  </div>

  <!-- Details 테이블 -->
  <div class="details-block">
    <table>
      <tr><td>Target release</td><td>{today}</td></tr>
      <tr><td>Project</td><td>{project}</td></tr>
      <tr><td>Document status</td><td><span class="badge-draft">DRAFT</span></td></tr>
      <tr><td>Document owner</td><td><span class="badge-owner">Requirements Analysis Agent</span></td></tr>
      <tr><td>Analysis mode</td><td>{mode or "통합 분석"}</td></tr>
      <tr><td>Quality Gate</td><td><strong style="color:{'#27ae60' if qc_pass == qc_total else '#e74c3c'}">{qc_pass}/{qc_total} PASS</strong></td></tr>
      <tr><td>Tech lead</td><td></td></tr>
      <tr><td>QA</td><td></td></tr>
    </table>
  </div>

  <hr class="divider">

  <!-- Objective -->
  <div class="section">
    <h2>Objective</h2>
    <p class="section-desc">Provide context on this project and explain how it fits into your organization's strategic goals</p>
    <p style="color:#333">
      본 문서는 <strong>{project}</strong> 프로젝트의 요구사항을 정의하고,
      이해관계자 간의 공통된 이해를 바탕으로 성공적인 시스템 구축을 지원합니다.
      총 <strong>{len(fr_rows)}개</strong>의 기능 요구사항과
      <strong>{len(nfr_rows)}개</strong>의 비기능 요구사항이 도출되었습니다.
    </p>
  </div>

  <hr class="divider">

  <!-- Success metrics -->
  <div class="section">
    <h2>Success metrics</h2>
    <p class="section-desc">List project goals and the metrics you'll use to judge its success</p>
    <table class="data-table">
      <thead><tr><th>Goal</th><th>Metric</th></tr></thead>
      <tbody>{kpi_rows_html}</tbody>
    </table>
  </div>

  <hr class="divider">

  <!-- Assumptions -->
  <div class="section">
    <h2>Assumptions</h2>
    <p class="section-desc">List any assumptions you have about your users, technical constraints, or business goals</p>
    <table class="data-table">
      <thead><tr><th>#</th><th>가정 사항</th><th>근거</th></tr></thead>
      <tbody>
        <tr><td class="num">1</td><td>사용자는 웹 브라우저 환경에서 시스템에 접근한다</td><td></td></tr>
        <tr><td class="num">2</td><td>회의록은 한국어로 작성된다</td><td></td></tr>
        <tr><td class="num">3</td><td>분석 대상 문서의 최대 크기는 20MB 이하이다</td><td></td></tr>
      </tbody>
    </table>
  </div>

  <hr class="divider">

  <!-- Milestones -->
  <div class="section">
    <h2>Milestones</h2>
    <p class="section-desc">Action items and target dates to keep your team on track</p>
    <table class="data-table">
      <thead><tr><th>#</th><th>액션 아이템</th><th>담당자</th><th>목표일</th><th>상태</th></tr></thead>
      <tbody>{act_rows_html}</tbody>
    </table>
  </div>

  <hr class="divider">

  <!-- Requirements (FR) -->
  <div class="section">
    <h2>Requirements — Functional (FR)</h2>
    <p class="section-desc">시스템이 반드시 수행해야 하는 기능적 요구사항 목록입니다.</p>
    <table class="data-table">
      <thead>
        <tr>
          <th class="num">#</th>
          <th>Requirement</th>
          <th>수용기준</th>
          <th>Importance</th>
          <th>Jira Issue</th>
          <th>Notes</th>
        </tr>
      </thead>
      <tbody>{fr_rows_html if fr_rows_html else '<tr><td colspan="6" style="color:#aaa;text-align:center">데이터 없음</td></tr>'}</tbody>
    </table>
  </div>

  <hr class="divider">

  <!-- Requirements (NFR) -->
  <div class="section">
    <h2>Requirements — Non-Functional (NFR)</h2>
    <p class="section-desc">성능·보안·가용성 등 시스템 품질 속성 요구사항입니다.</p>
    <table class="data-table">
      <thead>
        <tr>
          <th class="num">#</th>
          <th>Requirement</th>
          <th>User Story</th>
          <th>Importance</th>
          <th>Jira Issue</th>
          <th>Notes</th>
        </tr>
      </thead>
      <tbody>{nfr_rows_html if nfr_rows_html else '<tr><td colspan="6" style="color:#aaa;text-align:center">데이터 없음</td></tr>'}</tbody>
    </table>
  </div>

  <hr class="divider">

  <!-- User interaction and design -->
  <div class="section">
    <h2>User interaction and design</h2>
    <p class="section-desc">Add mockups, diagrams, or visual designs related to these requirements.</p>
    <div style="border:1px dashed #ddd;border-radius:6px;padding:24px;text-align:center;color:#bbb;font-size:12px;margin:10px 0">
      화면 설계 및 UI/UX 다이어그램을 여기에 첨부하세요.
    </div>
  </div>

  <hr class="divider">

  <!-- Open Questions -->
  <div class="section">
    <h2>Open Questions</h2>
    <table class="data-table">
      <thead><tr><th>Question</th><th>Answer</th><th>Date Answered</th></tr></thead>
      <tbody>
        <tr>
          <td style="color:#aaa">e.g., 추가 검토가 필요한 사항을 여기에 기록하세요.</td>
          <td></td><td></td>
        </tr>
      </tbody>
    </table>
  </div>

  <hr class="divider">

  <!-- Out of Scope -->
  <div class="section">
    <h2>Out of Scope</h2>
    <p class="section-desc">List the features discussed which are out of scope or might be revisited in a later release.</p>
    <table class="data-table">
      <thead><tr><th>#</th><th>항목</th><th>제외 사유</th></tr></thead>
      <tbody>
        <tr><td class="num">1</td>
          <td style="color:#aaa">MoSCoW 분석에서 Won't로 분류된 항목을 기록하세요.</td>
          <td></td>
        </tr>
      </tbody>
    </table>
  </div>

  <hr class="divider">

  <!-- Quality Gate -->
  <div class="section">
    <h2>Quality Gate</h2>
    <p class="section-desc">요구사항 분석의 완성도 및 품질 검증 결과입니다.</p>
    <table class="data-table">
      <thead><tr><th>검증 항목</th><th>결과</th><th>설명</th></tr></thead>
      <tbody>{qc_rows_html if qc_rows_html else '<tr><td colspan="3" style="color:#aaa">데이터 없음</td></tr>'}</tbody>
    </table>
  </div>

  <div style="margin-top:48px;padding-top:16px;border-top:1px solid #f0f0f0;font-size:11px;color:#bbb;text-align:center">
    Generated by Requirements Analysis Agent · {today} · {project}
  </div>

</div>

<script>
  // 페이지 로드 시 자동으로 인쇄 다이얼로그 열기
  window.addEventListener('load', function() {{
    setTimeout(function() {{ window.print(); }}, 800);
  }});
</script>
</body>
</html>"""
