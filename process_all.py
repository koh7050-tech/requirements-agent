#!/usr/bin/env python3
"""
Requirements Analysis Agent — Batch Runner
inputs/ 의 모든 CSV/XLSX → 수집→정의→분석→KPI→액션→추적성
outputs/YYYY-MM-DD/ 에 파일별 XLSX + Executive Summary + 품질Gate 저장
품질 Gate FAIL 항목은 확인_필요.md 로 별도 정리
"""
from __future__ import annotations

import csv
import json
import os
import sys
from datetime import date
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl 필요: pip3 install openpyxl")
    sys.exit(1)

try:
    import anthropic as _anthropic_mod
    _ANTHROPIC_AVAILABLE = True
except ImportError:
    _anthropic_mod = None
    _ANTHROPIC_AVAILABLE = False


# ── 스타일 상수 ───────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
MUST_FILL    = PatternFill("solid", fgColor="FFF2CC")
SHOULD_FILL  = PatternFill("solid", fgColor="E2EFDA")
COULD_FILL   = PatternFill("solid", fgColor="DDEBF7")
WONT_FILL    = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
FAIL_FILL    = PatternFill("solid", fgColor="FFE0E0")
PASS_FILL    = PatternFill("solid", fgColor="E0FFE0")

THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PRIORITY_FILL = {"Must": MUST_FILL, "Should": SHOULD_FILL,
                 "Could": COULD_FILL, "Won't": WONT_FILL}


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def _cell(ws, row, col, value, fill=None, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=10)
    c.fill = fill or WHITE_FILL
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = BORDER


def _set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, ref="A2"):
    ws.freeze_panes = ref


# ── 파일 로드 ────────────────────────────────────────────────────

def load_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_xlsx_as_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            result.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
    wb.close()
    return result


def load_input_file(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv(path)
    if suffix in (".xlsx", ".xls"):
        return load_xlsx_as_rows(path)
    raise ValueError(f"지원하지 않는 파일 형식: {suffix}")


def discover_inputs(folder: Path) -> list[Path]:
    files = []
    for pattern in ("*.csv", "*.CSV", "*.xlsx", "*.XLSX", "*.xls"):
        files.extend(folder.glob(pattern))
    seen = set()
    result = []
    for f in sorted(files):
        if f.name not in seen:
            seen.add(f.name)
            result.append(f)
    return result


# ── Claude API 파이프라인 ────────────────────────────────────────

def _pipeline_via_claude(rows: list[dict], api_key: str, file_name: str) -> dict:
    client = _anthropic_mod.Anthropic(api_key=api_key)

    sample_keys = list(rows[0].keys()) if rows else []
    text_rows = []
    for i, r in enumerate(rows):
        parts = [f"{k}: {v}" for k, v in r.items() if v and str(v).strip()]
        text_rows.append(f"[{i+1}] " + " | ".join(parts))
    interview_text = "\n".join(text_rows)

    prompt = f"""아래 인터뷰/회의록 데이터({file_name})에서 요구사항을 분석하여 JSON으로 반환하세요.

데이터:
{interview_text}

반환 형식 (JSON만, 마크다운 없이):
{{
  "raw_requirements": [
    {{"Raw_ID":"R-001","Source_ID":"INT-001","원문":"...","유형":"기능|비기능|제약|질문"}}
  ],
  "functional_requirements": [
    {{"ID":"FR-001","요구사항":"시스템은 ... 해야 한다","이해관계자":"...","우선순위":"Must|Should|Could|Won't","수용기준":"...","Raw_IDs":"R-001","Source_IDs":"INT-001"}}
  ],
  "nonfunctional_requirements": [
    {{"ID":"NFR-001","요구사항":"...","유형":"성능|보안|호환성|확장성|가용성|유지보수성","이해관계자":"...","우선순위":"Must|Should|Could|Won't","Raw_IDs":"R-002","Source_IDs":"INT-002"}}
  ],
  "moscow": [
    {{"ID":"FR-001","분류":"Must|Should|Could|Won't","복잡도":"상|중|하","리스크":"상|중|하","의존성":"없음|FR-002"}}
  ],
  "kpis": [
    {{"KPI_ID":"KPI-001","연결_REQ":"FR-001","목적":"...","지표명":"...","산식":"...","단위":"...","데이터원천":"...","측정주기":"월|주|일","기준값":"...","목표값":"...","목표기한":"YYYY-MM-DD","오너":"...","검증방법":"..."}}
  ],
  "actions": [
    {{"ACT_ID":"ACT-001","연결_REQ":"FR-001","액션":"...","오너":"...","목표일":"YYYY-MM-DD","상태":"대기|진행중|완료","우선순위":"Must|Should|Could"}}
  ],
  "traceability": [
    {{"Source_ID":"INT-001","Raw_ID":"R-001","REQ_ID":"FR-001","KPI_ID":"KPI-001","ACT_ID":"ACT-001","상태":"Active"}}
  ]
}}"""

    msg = client.messages.create(
        model="claude-opus-4-7",
        max_tokens=8192,
        thinking={"type": "adaptive"},
        messages=[{"role": "user", "content": prompt}],
    )
    text = next((b.text for b in msg.content if hasattr(b, "text")), "")
    start = text.find("{")
    end   = text.rfind("}") + 1
    if start == -1:
        raise ValueError("Claude API가 유효한 JSON을 반환하지 않았습니다.")
    return json.loads(text[start:end])


# ── 규칙 기반 파이프라인 ─────────────────────────────────────────

def _infer_source_id(row: dict, idx: int) -> str:
    for key in ("interview_id", "id", "ID", "source_id", "Source_ID"):
        if row.get(key, "").strip():
            return row[key].strip()
    return f"SRC-{idx:03d}"


def _infer_text(row: dict) -> str:
    for key in ("answer", "내용", "요구사항", "content", "text", "발언", "요구"):
        if row.get(key, "").strip():
            return row[key].strip()
    return " | ".join(str(v) for v in row.values() if str(v).strip())


def _infer_role(row: dict) -> str:
    for key in ("role", "역할", "담당자", "작성자", "이해관계자"):
        if row.get(key, "").strip():
            return row[key].strip()
    return "미지정"


_KEYWORD_FR: list[tuple[str, str, str, str, str]] = [
    ("인증",      "사용자 인증(로그인/로그아웃) 기능을 제공해야 한다",       "보안팀/사업기획",  "Must",   "로그인 성공 시 세션 발급, 로그아웃 시 세션 만료"),
    ("로그인",    "사용자 로그인/로그아웃 기능을 제공해야 한다",             "사용자/보안팀",    "Must",   "로그인 성공률 99% 이상, 실패 시 오류 메시지 표시"),
    ("대시보드",  "주요 지표를 시각화한 대시보드를 제공해야 한다",           "사업기획팀장",    "Should", "핵심 지표 5개 이상 시각화, 필터 기능 포함"),
    ("API 연동",  "기존 시스템과 REST API로 데이터를 연동해야 한다",         "개발팀장",        "Must",   "기존 API 명세 100% 호환, 연동 성공률 99% 이상"),
    ("알림",      "주요 이벤트 발생 시 사용자에게 알림을 제공해야 한다",     "사용자/운영팀",   "Should", "알림 발송 후 5분 내 도달, 발송 실패율 1% 미만"),
    ("검색",      "콘텐츠 및 데이터 검색 기능을 제공해야 한다",             "사용자",          "Should", "검색 결과 3초 이내 반환, 관련도 정렬 적용"),
    ("권한",      "역할 기반 접근 권한 관리 기능을 제공해야 한다",          "보안팀/관리자",   "Must",   "역할(Admin/User) 구분, 권한 외 접근 시 403 반환"),
    ("보고서",    "분석 결과를 보고서로 내보내기 할 수 있어야 한다",         "사업기획팀장",    "Could",  "PDF/Excel 형식 지원, 생성 30초 이내"),
]

_KEYWORD_NFR: list[tuple[str, str, str, str]] = [
    ("API",       "시스템은 기존 API 스펙을 준수해야 한다",          "호환성",   "Must"),
    ("성능",      "시스템 응답시간은 3초 이내여야 한다",             "성능",     "Must"),
    ("보안",      "전송 데이터는 TLS 1.2 이상으로 암호화해야 한다", "보안",     "Must"),
    ("가용성",    "시스템 가용성은 99.9% 이상이어야 한다",           "가용성",   "Should"),
    ("확장성",    "트래픽 증가에 따라 수평 확장이 가능해야 한다",    "확장성",   "Should"),
]


def _pipeline_rule_based(rows: list[dict], file_name: str) -> dict:
    combined = " ".join(_infer_text(r) for r in rows)

    raw = [
        {"Raw_ID": f"R-{i:03d}", "Source_ID": _infer_source_id(r, i),
         "원문": _infer_text(r), "유형": "기능"}
        for i, r in enumerate(rows, 1)
    ]

    frs, nfrs = [], []
    fr_id = nfr_id = 1
    used_fr: set[str] = set()
    used_nfr: set[str] = set()

    # 원문에서 출처 매핑
    def find_source(kw: str) -> tuple[str, str]:
        for i, r in enumerate(rows, 1):
            if kw in _infer_text(r):
                return f"R-{i:03d}", _infer_source_id(r, i)
        return "R-001", raw[0]["Source_ID"] if raw else "SRC-001"

    for kw, req, stakeholder, priority, criteria in _KEYWORD_FR:
        if kw in combined and kw not in used_fr:
            raw_id, src_id = find_source(kw)
            frs.append({"ID": f"FR-{fr_id:03d}", "요구사항": req, "이해관계자": stakeholder,
                        "우선순위": priority, "수용기준": criteria,
                        "Raw_IDs": raw_id, "Source_IDs": src_id})
            fr_id += 1
            used_fr.add(kw)

    for kw, req, req_type, priority in _KEYWORD_NFR:
        if kw in combined and kw not in used_nfr:
            raw_id, src_id = find_source(kw)
            role = next((_infer_role(r) for r in rows if kw in _infer_text(r)), "개발팀장")
            nfrs.append({"ID": f"NFR-{nfr_id:03d}", "요구사항": req, "유형": req_type,
                         "이해관계자": role, "우선순위": priority,
                         "Raw_IDs": raw_id, "Source_IDs": src_id})
            nfr_id += 1
            used_nfr.add(kw)

    # FR/NFR 없으면 원문 기반 최소 생성
    if not frs:
        for i, r in enumerate(rows, 1):
            t = _infer_text(r)
            if t:
                frs.append({"ID": f"FR-{i:03d}",
                            "요구사항": f"시스템은 다음을 지원해야 한다: {t[:80]}",
                            "이해관계자": _infer_role(r), "우선순위": "Should",
                            "수용기준": "담당자 확인 필요",
                            "Raw_IDs": f"R-{i:03d}", "Source_IDs": _infer_source_id(r, i)})

    moscow = [
        {"ID": fr["ID"], "분류": fr["우선순위"], "복잡도": "중", "리스크": "중", "의존성": "없음"}
        for fr in frs
    ] + [
        {"ID": nfr["ID"], "분류": nfr["우선순위"], "복잡도": "하", "리스크": "하", "의존성": "없음"}
        for nfr in nfrs
    ]

    _KPI_TEMPLATE = {
        "Must":   ("99% 이상", "월"),
        "Should": ("95% 이상", "월"),
        "Could":  ("목표 달성", "분기"),
    }
    kpis = []
    kpi_id = 1
    for fr in frs:
        target, cycle = _KPI_TEMPLATE.get(fr["우선순위"], ("목표 달성", "월"))
        kpis.append({
            "KPI_ID": f"KPI-{kpi_id:03d}", "연결_REQ": fr["ID"],
            "목적": f"{fr['요구사항'][:30]} 목표 달성",
            "지표명": f"{fr['ID']} 완료율",
            "산식": "완료 항목 수 / 전체 항목 수 × 100",
            "단위": "%", "데이터원천": "시스템 로그/APM",
            "측정주기": cycle, "기준값": "N/A", "목표값": target,
            "목표기한": "2026-12-31", "오너": fr["이해관계자"],
            "검증방법": "정기 리뷰 미팅",
        })
        kpi_id += 1

    _ACT_DEADLINE = {"Must": "2026-08-31", "Should": "2026-09-30", "Could": "2026-10-31"}
    actions = []
    act_id = 1
    for req in frs + nfrs:
        actions.append({
            "ACT_ID": f"ACT-{act_id:03d}", "연결_REQ": req["ID"],
            "액션": f"{req['ID']} 구현 및 검증",
            "오너": req["이해관계자"],
            "목표일": _ACT_DEADLINE.get(req["우선순위"], "2026-12-31"),
            "상태": "대기", "우선순위": req["우선순위"],
        })
        act_id += 1

    traceability = []
    for r in raw:
        matched = [fr for fr in frs if fr["Raw_IDs"] == r["Raw_ID"]]
        if not matched:
            continue
        for fr in matched:
            kpi = next((k for k in kpis if k["연결_REQ"] == fr["ID"]), None)
            act = next((a for a in actions if a["연결_REQ"] == fr["ID"]), None)
            traceability.append({
                "Source_ID": r["Source_ID"], "Raw_ID": r["Raw_ID"],
                "REQ_ID": fr["ID"],
                "KPI_ID": kpi["KPI_ID"] if kpi else "-",
                "ACT_ID": act["ACT_ID"] if act else "-",
                "상태": "Active",
            })

    return {
        "raw_requirements": raw,
        "functional_requirements": frs,
        "nonfunctional_requirements": nfrs,
        "moscow": moscow,
        "kpis": kpis,
        "actions": actions,
        "traceability": traceability,
    }


# ── 파이프라인 진입점 ─────────────────────────────────────────────

def build_pipeline_data(rows: list[dict], file_name: str) -> dict:
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if api_key and _ANTHROPIC_AVAILABLE:
        print("      [Claude API 모드] 분석 중...")
        try:
            return _pipeline_via_claude(rows, api_key, file_name)
        except Exception as e:
            print(f"      [경고] Claude API 실패 ({e}) → 규칙 기반 전환")
    else:
        print("      [규칙 기반 모드] 분석 중...")
    return _pipeline_rule_based(rows, file_name)


# ── 품질 Gate ─────────────────────────────────────────────────────

def run_quality_gate(data: dict, file_name: str) -> list[dict]:
    """
    품질 Gate 항목 리스트 반환.
    각 항목: {file, check, passed, detail}
    """
    frs  = data["functional_requirements"]
    nfrs = data["nonfunctional_requirements"]
    all_reqs = frs + nfrs
    kpis = data["kpis"]
    acts = data["actions"]
    trace = data["traceability"]

    checks = []

    # 1. 모든 요구사항에 출처 존재
    missing_src = [r["ID"] for r in all_reqs if not r.get("Source_IDs", "").strip()]
    checks.append({
        "file": file_name, "check": "모든 요구사항에 출처(Source_IDs) 존재",
        "passed": len(missing_src) == 0,
        "detail": f"출처 누락: {', '.join(missing_src)}" if missing_src else "-",
    })

    # 2. Must 요구사항에 수용기준 존재
    must_no_criteria = [r["ID"] for r in all_reqs
                        if r.get("우선순위") == "Must" and not r.get("수용기준", "").strip()]
    checks.append({
        "file": file_name, "check": "Must 요구사항에 수용기준 존재",
        "passed": len(must_no_criteria) == 0,
        "detail": f"수용기준 누락: {', '.join(must_no_criteria)}" if must_no_criteria else "-",
    })

    # 3. 모든 FR에 KPI 연결
    fr_ids = {fr["ID"] for fr in frs}
    kpi_linked = {k["연결_REQ"] for k in kpis}
    fr_no_kpi = fr_ids - kpi_linked
    checks.append({
        "file": file_name, "check": "모든 FR에 KPI 연결",
        "passed": len(fr_no_kpi) == 0,
        "detail": f"KPI 미연결 FR: {', '.join(sorted(fr_no_kpi))}" if fr_no_kpi else "-",
    })

    # 4. KPI 산식과 데이터 원천 존재
    kpi_incomplete = [k["KPI_ID"] for k in kpis if not k.get("산식") or not k.get("데이터원천")]
    checks.append({
        "file": file_name, "check": "KPI 산식 및 데이터 원천 존재",
        "passed": len(kpi_incomplete) == 0,
        "detail": f"불완전 KPI: {', '.join(kpi_incomplete)}" if kpi_incomplete else "-",
    })

    # 5. 액션 오너와 목표일 존재
    act_incomplete = [a["ACT_ID"] for a in acts if not a.get("오너") or not a.get("목표일")]
    checks.append({
        "file": file_name, "check": "액션 오너와 목표일 존재",
        "passed": len(act_incomplete) == 0,
        "detail": f"불완전 액션: {', '.join(act_incomplete)}" if act_incomplete else "-",
    })

    # 6. Must 요구사항에 액션 연결
    must_ids = {r["ID"] for r in all_reqs if r.get("우선순위") == "Must"}
    act_linked = {a["연결_REQ"] for a in acts}
    must_no_act = must_ids - act_linked
    checks.append({
        "file": file_name, "check": "Must 요구사항에 액션 연결",
        "passed": len(must_no_act) == 0,
        "detail": f"액션 미연결 Must: {', '.join(sorted(must_no_act))}" if must_no_act else "-",
    })

    # 7. 추적성 매트릭스 커버리지 50% 이상
    coverage = len(trace) / max(len(frs), 1)
    checks.append({
        "file": file_name, "check": "추적성 커버리지 50% 이상 (FR 기준)",
        "passed": coverage >= 0.5,
        "detail": f"커버리지 {coverage*100:.0f}% ({len(trace)}/{len(frs)})" if not coverage >= 0.5 else "-",
    })

    # 8. 요구사항 수 최소 1건 이상
    checks.append({
        "file": file_name, "check": "요구사항 최소 1건 이상 존재",
        "passed": len(all_reqs) > 0,
        "detail": "요구사항 없음" if len(all_reqs) == 0 else "-",
    })

    return checks


# ── XLSX 시트 생성 ────────────────────────────────────────────────

def _sheet_source(wb, rows: list[dict]):
    ws = wb.create_sheet("1_출처")
    if not rows:
        return
    headers = list(rows[0].keys())
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            _cell(ws, r, c, row.get(h, ""))
    _set_widths(ws, [max(12, min(40, len(h) + 4)) for h in headers])
    _freeze(ws)


def _sheet_raw(wb, data: list[dict]):
    ws = wb.create_sheet("2_원문요구")
    headers = ["Raw_ID", "Source_ID", "원문", "유형"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    for r, row in enumerate(data, 2):
        for c, h in enumerate(headers, 1):
            _cell(ws, r, c, row.get(h, ""))
    _set_widths(ws, [10, 12, 55, 12])
    _freeze(ws)


def _sheet_fr(wb, frs: list[dict]):
    ws = wb.create_sheet("3_기능요구사항")
    headers = ["ID", "요구사항", "이해관계자", "우선순위", "수용기준", "Raw_IDs", "Source_IDs", "상태"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    for r, row in enumerate(frs, 2):
        fill = PRIORITY_FILL.get(row.get("우선순위", ""), WHITE_FILL)
        for c, h in enumerate(headers, 1):
            _cell(ws, r, c, row.get(h, "Active" if h == "상태" else ""), fill=fill)
    _set_widths(ws, [10, 45, 16, 10, 35, 10, 12, 10])
    _freeze(ws)


def _sheet_nfr(wb, nfrs: list[dict]):
    ws = wb.create_sheet("4_비기능요구사항")
    headers = ["ID", "요구사항", "유형", "이해관계자", "우선순위", "Raw_IDs", "Source_IDs", "상태"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    for r, row in enumerate(nfrs, 2):
        fill = PRIORITY_FILL.get(row.get("우선순위", ""), WHITE_FILL)
        for c, h in enumerate(headers, 1):
            _cell(ws, r, c, row.get(h, "Active" if h == "상태" else ""), fill=fill)
    _set_widths(ws, [10, 45, 12, 16, 10, 10, 12, 10])
    _freeze(ws)


def _sheet_moscow(wb, moscow: list[dict]):
    ws = wb.create_sheet("5_분석_MoSCoW")
    headers = ["ID", "분류", "복잡도", "리스크", "의존성"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    for r, row in enumerate(moscow, 2):
        fill = PRIORITY_FILL.get(row.get("분류", ""), WHITE_FILL)
        for c, h in enumerate(headers, 1):
            _cell(ws, r, c, row.get(h, ""), fill=fill)
    _set_widths(ws, [10, 10, 10, 10, 22])
    _freeze(ws)


def _sheet_kpi(wb, kpis: list[dict]):
    ws = wb.create_sheet("6_KPI")
    headers = ["KPI_ID", "연결_REQ", "목적", "지표명", "산식", "단위",
               "데이터원천", "측정주기", "기준값", "목표값", "목표기한", "오너", "검증방법"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    for r, row in enumerate(kpis, 2):
        for c, h in enumerate(headers, 1):
            _cell(ws, r, c, row.get(h, ""))
    _set_widths(ws, [10, 10, 22, 22, 38, 8, 22, 10, 10, 12, 14, 16, 26])
    _freeze(ws)


def _sheet_actions(wb, actions: list[dict]):
    ws = wb.create_sheet("7_액션아이템")
    headers = ["ACT_ID", "연결_REQ", "액션", "오너", "목표일", "상태", "우선순위"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    for r, row in enumerate(actions, 2):
        fill = PRIORITY_FILL.get(row.get("우선순위", ""), WHITE_FILL)
        for c, h in enumerate(headers, 1):
            _cell(ws, r, c, row.get(h, ""), fill=fill)
    _set_widths(ws, [10, 10, 42, 16, 14, 10, 10])
    _freeze(ws)


def _sheet_traceability(wb, trace: list[dict]):
    ws = wb.create_sheet("8_추적성")
    headers = ["Source_ID", "Raw_ID", "REQ_ID", "KPI_ID", "ACT_ID", "상태"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    for r, row in enumerate(trace, 2):
        for c, h in enumerate(headers, 1):
            _cell(ws, r, c, row.get(h, ""), align="center")
    _set_widths(ws, [14, 10, 10, 10, 10, 10])
    _freeze(ws)


def _sheet_quality(wb, checks: list[dict]):
    ws = wb.create_sheet("9_품질Gate")
    _hdr(ws, 1, 1, "점검 항목")
    _hdr(ws, 1, 2, "결과")
    _hdr(ws, 1, 3, "상세")
    for r, ch in enumerate(checks, 2):
        _cell(ws, r, 1, ch["check"])
        passed = ch["passed"]
        result_cell = ws.cell(row=r, column=2, value="PASS" if passed else "FAIL")
        result_cell.font = Font(bold=True, color="00B050" if passed else "FF0000", size=10)
        result_cell.fill = PASS_FILL if passed else FAIL_FILL
        result_cell.alignment = Alignment(horizontal="center", vertical="center")
        result_cell.border = BORDER
        _cell(ws, r, 3, ch.get("detail", "-"),
              fill=FAIL_FILL if not passed else WHITE_FILL)
    _set_widths(ws, [42, 10, 45])
    _freeze(ws)


def build_xlsx(source_rows: list[dict], data: dict, checks: list[dict], out_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_source(wb, source_rows)
    _sheet_raw(wb, data["raw_requirements"])
    _sheet_fr(wb, data["functional_requirements"])
    _sheet_nfr(wb, data["nonfunctional_requirements"])
    _sheet_moscow(wb, data["moscow"])
    _sheet_kpi(wb, data["kpis"])
    _sheet_actions(wb, data["actions"])
    _sheet_traceability(wb, data["traceability"])
    _sheet_quality(wb, checks)
    wb.save(out_path)


# ── Executive Summary ─────────────────────────────────────────────

def generate_summary(data: dict, project: str, source_rows: list[dict],
                     checks: list[dict], file_name: str) -> str:
    today = date.today().isoformat()
    frs   = data["functional_requirements"]
    nfrs  = data["nonfunctional_requirements"]
    kpis  = data["kpis"]
    acts  = data["actions"]
    trace = data["traceability"]
    must_acts = [a for a in acts if a.get("우선순위") == "Must"]
    failed = [c for c in checks if not c["passed"]]
    passed_count = sum(1 for c in checks if c["passed"])

    fr_lines  = "\n".join(f"  - {r['ID']}: {r['요구사항']} [{r['우선순위']}]" for r in frs)
    nfr_lines = "\n".join(f"  - {r['ID']}: {r['요구사항']} [{r.get('유형','')}] [{r['우선순위']}]" for r in nfrs)
    kpi_lines = "\n".join(f"  - {k['KPI_ID']}: {k['지표명']} — 목표 {k['목표값']} ({k['측정주기']})" for k in kpis)
    act_lines = "\n".join(f"  - {a['ACT_ID']}: {a['액션']} | {a['오너']} | {a['목표일']}" for a in must_acts)
    fail_lines = "\n".join(f"  - [{c['check']}] {c['detail']}" for c in failed) or "  - 없음 (전체 PASS)"

    return f"""# Executive Summary — 요구사항 분석

| 항목 | 내용 |
|---|---|
| 프로젝트 | {project} |
| 원본 파일 | {file_name} |
| 작성일 | {today} |
| 출처 수 | {len(source_rows)}건 |
| 기능 요구사항 | {len(frs)}건 |
| 비기능 요구사항 | {len(nfrs)}건 |
| KPI | {len(kpis)}건 |
| 액션 아이템 | {len(acts)}건 |
| 추적성 연결 | {len(trace)}건 |
| 품질 Gate | {passed_count}/{len(checks)} PASS |

---

## 기능 요구사항 ({len(frs)}건)

{fr_lines if fr_lines else "  - 없음"}

## 비기능 요구사항 ({len(nfrs)}건)

{nfr_lines if nfr_lines else "  - 없음"}

---

## KPI 요약 ({len(kpis)}건)

{kpi_lines if kpi_lines else "  - 없음"}

---

## Must 액션 아이템 ({len(must_acts)}건)

{act_lines if act_lines else "  - 없음"}

---

## 품질 Gate 결과 ({passed_count}/{len(checks)} PASS)

{fail_lines}

---
*본 문서는 Requirements Analysis Agent에 의해 자동 생성되었습니다. ({today})*
"""


# ── 확인 필요 보고서 ──────────────────────────────────────────────

def generate_review_report(all_checks: list[dict]) -> str:
    today = date.today().isoformat()
    failed = [c for c in all_checks if not c["passed"]]
    total = len(all_checks)
    fail_count = len(failed)

    if not failed:
        body = "모든 품질 Gate 항목이 통과되었습니다."
    else:
        lines = []
        by_file: dict[str, list[dict]] = {}
        for c in failed:
            by_file.setdefault(c["file"], []).append(c)
        for fname, items in by_file.items():
            lines.append(f"\n### {fname}")
            for item in items:
                lines.append(f"- **[FAIL]** {item['check']}")
                if item.get("detail") and item["detail"] != "-":
                    lines.append(f"  - 상세: {item['detail']}")
        body = "\n".join(lines)

    return f"""# 확인 필요 항목 — 품질 Gate FAIL 보고서

| 항목 | 내용 |
|---|---|
| 작성일 | {today} |
| 전체 점검 수 | {total}건 |
| PASS | {total - fail_count}건 |
| **FAIL** | **{fail_count}건** |

---

## FAIL 항목 목록

{body}

---

## 권고 조치

| 순위 | 항목 | 조치 |
|---|---|---|
| 1 | 수용기준 누락 | Must 요구사항 담당자와 협의하여 수용기준 작성 |
| 2 | 출처 누락 | 원본 인터뷰/회의록 재확인 후 Source_IDs 보완 |
| 3 | KPI 미연결 | 해당 FR 담당자와 KPI 지표 정의 미팅 진행 |
| 4 | 추적성 커버리지 부족 | 추적성 매트릭스 작성 워크숍 진행 |

---
*본 문서는 Requirements Analysis Agent — 품질 Gate 모듈에 의해 자동 생성되었습니다. ({today})*
"""


# ── 메인 ─────────────────────────────────────────────────────────

def main():
    base_dir   = Path(__file__).parent
    inputs_dir = base_dir / "inputs"
    today_str  = date.today().isoformat()
    out_dir    = base_dir / "outputs" / today_str
    out_dir.mkdir(parents=True, exist_ok=True)

    input_files = discover_inputs(inputs_dir)
    if not input_files:
        print(f"[ERROR] inputs/ 에 CSV/XLSX 파일이 없습니다: {inputs_dir}")
        sys.exit(1)

    print(f"[발견] {len(input_files)}개 파일: {[f.name for f in input_files]}")
    print(f"[출력] {out_dir}\n")

    all_checks: list[dict] = []

    for path in input_files:
        print(f"{'='*55}")
        print(f"  처리 중: {path.name}")
        print(f"{'='*55}")

        # 1. 로드
        print("[1/4] 파일 로드...")
        rows = load_input_file(path)
        print(f"      → {len(rows)}행")

        # 2. 파이프라인
        print("[2/4] 수집→정의→분석→KPI→액션→추적성 파이프라인...")
        data = build_pipeline_data(rows, path.name)
        frs  = data["functional_requirements"]
        nfrs = data["nonfunctional_requirements"]
        print(f"      → FR {len(frs)}건, NFR {len(nfrs)}건")
        print(f"      → KPI {len(data['kpis'])}건, 액션 {len(data['actions'])}건, 추적성 {len(data['traceability'])}건")

        # 3. 품질 Gate
        print("[3/4] 품질 Gate 점검...")
        checks = run_quality_gate(data, path.name)
        all_checks.extend(checks)
        pass_n = sum(1 for c in checks if c["passed"])
        print(f"      → {pass_n}/{len(checks)} PASS")
        for c in checks:
            status = "PASS" if c["passed"] else "FAIL"
            print(f"         [{status}] {c['check']}")
            if not c["passed"] and c.get("detail") != "-":
                print(f"                → {c['detail']}")

        # 4. 출력
        print("[4/4] 파일 저장...")
        stem = path.stem
        project = stem

        xlsx_path    = out_dir / f"{stem}_요구분석.xlsx"
        summary_path = out_dir / f"{stem}_Executive_Summary.md"

        build_xlsx(rows, data, checks, xlsx_path)
        summary = generate_summary(data, project, rows, checks, path.name)
        summary_path.write_text(summary, encoding="utf-8")

        print(f"      → XLSX    : {xlsx_path.name}")
        print(f"      → Summary : {summary_path.name}")
        print()

    # 통합 확인 필요 보고서
    review_path = out_dir / "확인_필요.md"
    review_text = generate_review_report(all_checks)
    review_path.write_text(review_text, encoding="utf-8")

    total = len(all_checks)
    fail_n = sum(1 for c in all_checks if not c["passed"])
    print(f"{'='*55}")
    print(f"✓ 전체 완료")
    print(f"  처리 파일 : {len(input_files)}개")
    print(f"  품질 Gate : {total - fail_n}/{total} PASS  ({fail_n}건 FAIL)")
    print(f"  출력 폴더 : {out_dir}")
    print(f"  확인 필요 : {review_path.name}")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
