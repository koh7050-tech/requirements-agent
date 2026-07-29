#!/usr/bin/env python3
"""
회의록 CSV → 요구분석 전체 파이프라인
수집→정의→분석→KPI→액션→추적성 + XLSX + Executive Summary
"""

from __future__ import annotations
import csv
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl 필요: pip3 install openpyxl")
    sys.exit(1)

try:
    import anthropic
except ImportError:
    anthropic = None


# ── 스타일 헬퍼 ──────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
MUST_FILL    = PatternFill("solid", fgColor="FFF2CC")
SHOULD_FILL  = PatternFill("solid", fgColor="E2EFDA")
COULD_FILL   = PatternFill("solid", fgColor="DDEBF7")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr(ws, row, col, value, bold=True, fg="FFFFFF", fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=10)
    cell.fill = fill or HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell

def cell(ws, row, col, value, fill=None, bold=False, wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=10)
    c.fill = fill or WHITE_FILL
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = BORDER
    return c

def set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, ref="A2"):
    ws.freeze_panes = ref


# ── CSV 로드 ─────────────────────────────────────────────────────

def load_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# ── 파이프라인 데이터 (Claude API 또는 내장 분석) ────────────────

def build_pipeline_data(rows: list[dict]) -> dict:
    """
    CSV 인터뷰 행에서 요구분석 전체 데이터를 구성.
    ANTHROPIC_API_KEY가 있으면 Claude로 생성, 없으면 규칙 기반 처리.
    """
    import os
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")

    if api_key and anthropic:
        try:
            return _pipeline_via_claude(rows, api_key)
        except Exception as e:
            print(f"[Claude API 실패] {e} → 규칙 기반으로 전환")
            return _pipeline_rule_based(rows)
    else:
        return _pipeline_rule_based(rows)


def _pipeline_via_claude(rows: list[dict], api_key: str) -> dict:
    """Claude API로 전체 파이프라인 실행 후 구조화 데이터 반환"""
    client = anthropic.Anthropic(api_key=api_key)

    interview_text = "\n".join(
        f"[{r['interview_id']}] {r['date']} | {r['role']} | Q: {r['question']} | A: {r['answer']}"
        for r in rows
    )

    prompt = f"""아래 인터뷰에서 요구사항을 분석하여 JSON으로 반환하세요.

인터뷰:
{interview_text}

반환 형식 (JSON만, 설명 없이):
{{
  "raw_requirements": [
    {{"Raw_ID":"R-001","Source_ID":"INT-001","원문":"...","유형":"기능|비기능|제약|질문"}}
  ],
  "functional_requirements": [
    {{"ID":"FR-001","요구사항":"시스템은 ... 해야 한다","이해관계자":"...","우선순위":"Must|Should|Could|Won't","수용기준":"...","Raw_IDs":"R-001","Source_IDs":"INT-001"}}
  ],
  "nonfunctional_requirements": [
    {{"ID":"NFR-001","요구사항":"...","유형":"성능|보안|호환성|확장성","이해관계자":"...","우선순위":"Must|Should|Could|Won't","수용기준":"...","Raw_IDs":"R-002","Source_IDs":"INT-002"}}
  ],
  "moscow": [
    {{"ID":"FR-001","분류":"Must","복잡도":"상|중|하","리스크":"상|중|하","의존성":"없음|FR-002"}}
  ],
  "kpis": [
    {{"KPI_ID":"KPI-001","연결_REQ":"FR-001","목적":"...","지표명":"...","산식":"...","단위":"...","데이터원천":"...","측정주기":"월|주|일","기준값":"...","목표값":"...","목표기한":"YYYY-MM-DD","오너":"...","검증방법":"..."}}
  ],
  "actions": [
    {{"ACT_ID":"ACT-001","연결_REQ":"FR-001","액션":"...","오너":"...","목표일":"YYYY-MM-DD","상태":"대기|진행중|완료","우선순위":"Must|Should|Could"}}
  ],
  "traceability": [
    {{"Source_ID":"INT-001","Raw_ID":"R-001","REQ_ID":"FR-001","KPI_ID":"KPI-001","ACT_ID":"ACT-001","상태":"Active"}}
  ]
}}"""

    import json
    msg = client.messages.create(
        model="claude-opus-4-7",
        max_tokens=4096,
        thinking={"type": "adaptive"},
        messages=[{"role": "user", "content": prompt}],
    )
    text = next((b.text for b in msg.content if hasattr(b, "text")), "")
    start = text.find("{")
    end   = text.rfind("}") + 1
    return json.loads(text[start:end])


def _pipeline_rule_based(rows: list[dict]) -> dict:
    """규칙 기반 파이프라인 — 인터뷰 행별 토픽 패턴으로 FR/NFR 추출"""
    # ── 원문 요구 목록 ───────────────────────────────────────────
    raw = []
    for i, r in enumerate(rows, 1):
        answer = r.get("answer", "")
        # 유형 분류: 인프라/보안/성능 키워드 → 비기능, 나머지 → 기능
        nfr_signals = ["이중화", "서버", "인프라", "SCP", "보안", "모니터링",
                       "성능", "부하", "배치", "주기", "로그", "리포트"]
        유형 = "비기능" if any(k in answer for k in nfr_signals) else "기능"
        raw.append({
            "Raw_ID": f"R-{i:03d}",
            "Source_ID": r.get("interview_id", f"INT-{i:03d}"),
            "원문": answer[:120] + ("..." if len(answer) > 120 else ""),
            "유형": 유형,
        })

    # ── FR 정의 테이블 (토픽 키워드 → 요구사항) ─────────────────
    # (키워드들, 요구사항문장, 이해관계자, 우선순위, 수용기준, source_id)
    FR_RULES = [
        (["공지사항", "공지"],
         "관리자가 공지사항을 등록하고 사용자가 조회할 수 있어야 한다",
         "제일기획_담당자", "Must",
         "관리자 등록 즉시 사용자 화면 노출, 조회수 집계",
         "INT-024"),
        (["FAQ"],
         "관리자가 FAQ 질문·답변을 등록하고 사용자가 펼침 방식으로 조회할 수 있어야 한다",
         "제일기획_담당자", "Must",
         "질문·답변 CRUD 가능, 사용자 화면 accordion 방식 표시",
         "INT-024"),
        (["VOC"],
         "사용자가 VOC를 등록하고 상태(접수·검토중·답변완료)를 추적할 수 있어야 한다",
         "제일기획_담당자", "Must",
         "VOC 등록 시 접수상태 자동부여, 답변 등록 시 답변완료 자동전환",
         "INT-024"),
        (["체크아웃", "체크인", "버전"],
         "체크아웃·체크인 시에만 버전이 증가하고 파일명에 날짜·시간이 자동부여되어야 한다",
         "제일기획_담당자", "Must",
         "체크인 완료 시 버전 +1, 파일명 뒤 YYYYMMDD_HHmmss 자동 부여",
         "INT-006"),
        (["라이브러리", "워킹"],
         "워킹 폴더의 최종 결과물만 라이브러리로 복사 등록되어야 한다",
         "제일기획_담당자", "Must",
         "워킹 원본 유지 후 라이브러리 복사, 60일 자동삭제 정책 적용",
         "INT-007"),
        (["폴더", "권한", "오너"],
         "워킹 프로젝트 생성자가 오너로서 폴더와 작업자를 관리할 수 있어야 한다",
         "제일기획_담당자", "Must",
         "오너만 최상위 폴더 생성·삭제 가능, 권한 없는 사용자는 폴더 비노출",
         "INT-008"),
        (["엠바고"],
         "라이브러리 자산에 엠바고 표시값을 부여하고 시각적 UI 경고를 제공해야 한다",
         "제일기획_담당자", "Must",
         "엠바고 여부 메타데이터 필드 존재, 폴더 라벨·배지·색상 UI 구현",
         "INT-011"),
        (["Knox", "로그인", "사용자 관리"],
         "Knox 계정 보유 임직원이 DAM에 로그인하고 운영자 승인 후 권한을 부여받아야 한다",
         "제일기획_담당자", "Must",
         "Knox 이메일 = DAM ID 동일 적용, 최초 로그인 시 비밀번호 변경 강제",
         "INT-012"),
        (["사용자 검색", "한글명", "영문명"],
         "사용자 ID·한글명·영문명·이메일로 동시 검색이 가능해야 한다",
         "제일기획_담당자", "Should",
         "시스템 언어 무관 한글명·영문명 동시 검색, 결과 즉시 노출",
         "INT-013"),
        (["워터마크"],
         "라이브러리 자산 프리뷰에 워터마크를 표시하고 원본 다운로드 시 미포함해야 한다",
         "제일기획_담당자", "Should",
         "프리뷰 화면 워터마크 100% 표시, 다운로드 원본에 워터마크 미삽입",
         "INT-014"),
        (["AI 분석", "KD", "LLM"],
         "라이브러리 등록 자산을 KD 1차·LLM 2차 순으로 AI 분석하고 결과를 메타데이터로 저장해야 한다",
         "펜타PM", "Must",
         "에셋 ID 기준 분석결과 저장, 중복분석 방지 로직 포함, 2000자 이내",
         "INT-016"),
        (["AI 태깅", "태깅", "자연어 검색"],
         "AI 태깅 결과로 자연어 검색이 가능하고 광고 목적·크리에이티브 맥락 검색을 지원해야 한다",
         "제일기획_담당자", "Must",
         "자연어 검색 정확도 목표치 달성, Visual·Business·Creative Context 항목 구분",
         "INT-018"),
        (["영상 분석", "영상"],
         "15초·30초 광고영상을 대표 프레임 추출 방식으로 AI 분석해야 한다",
         "제일기획_담당자", "Should",
         "대표 프레임 5~10장 추출, Gemma 기반 분석 가능범위 테스트 완료",
         "INT-019"),
        (["메타데이터"],
         "에셋에 기본정보·업무정보 메타데이터를 자동상속하고 AI 분석결과를 자동입력해야 한다",
         "제일기획_담당자", "Must",
         "상위 폴더 고객사·프로젝트 정보 자산 자동상속, 수동입력 항목 최소화",
         "INT-020"),
        (["AI 메타데이터 화면", "핵심 한 줄"],
         "AI 분석 결과를 핵심 한 줄 요약과 항목별 상세 필드로 구분하여 화면에 노출해야 한다",
         "제일기획_담당자", "Should",
         "LLM 한 줄 요약 상단 배치, Visual·Business·Creative 항목별 분리 노출",
         "INT-021"),
        (["원본파일", "메타데이터 삽입", "XML"],
         "다운로드 시점에 AI 태그를 원본 이미지·영상 파일에 삽입하거나 XML로 첨부할 수 있어야 한다",
         "제일기획_담당자", "Could",
         "JPG·PNG 파일 내 태그 삽입 가능, XML 별도 메타데이터 파일 첨부 선택 가능",
         "INT-022"),
        (["API", "외부시스템", "생성형"],
         "에셋 ID·프리뷰 URL 기반으로 외부시스템 연동 API를 제공해야 한다",
         "제일기획_담당자", "Could",
         "에셋 ID·프리뷰 URL API 응답 규격 정의, 자연어 검색 API 추가개발 포함",
         "INT-023"),
        (["요구사항 관리", "추적표", "196"],
         "기존 196개 요구사항 원본을 유지하고 변경·제외·추가 이력을 통합 추적표로 관리해야 한다",
         "제일기획_담당자", "Must",
         "제외 항목 삭제 금지, 사유·처리방향 표기, 제일기획·펜타 추적표 통합",
         "INT-003"),
        (["외부 협업", "Hub 폴더"],
         "외부 협력업체 협업은 Hub 폴더·외부 서버로 분리하고 관리자 승인형 권한구조로 운영해야 한다",
         "제일기획_담당자", "Could",
         "내부 워킹 폴더 직접 접근 차단, 외부 Hub 폴더 별도 권한구조 설계",
         "INT-009"),
        (["폴더 체계", "Workspace", "HQ"],
         "Workspace > HQ > Working Folder / Library 계층으로 폴더 체계를 구성해야 한다",
         "펜타PM", "Must",
         "폴더 생성 위치별 선택 가능한 폴더유형 제한, 고객사·프로젝트 폴더 자동 생성",
         "INT-010"),
    ]

    # ── NFR 정의 테이블 ──────────────────────────────────────────
    NFR_RULES = [
        (["DB 이중화", "Active-Standby"],
         "DB는 Active-Standby 이중화 구성으로 운영되어야 한다",
         "가용성", "제일기획_담당자", "Must",
         "Active 장애 시 Standby 자동 전환 30초 이내",
         "INT-005"),
        (["SCP", "Dell 서버", "수급"],
         "SCP 개발서버 및 Dell 서버 수급일정이 프로젝트 일정 내 확보되어야 한다",
         "인프라", "제일기획_담당자", "Must",
         "개발서버 수급 완료 후 개발 착수, 지연 시 대체 방안 수립",
         "INT-005"),
        (["대량 다운로드", "이상탐지", "모니터링"],
         "대량 다운로드 이상징후를 탐지하고 일 1회 리포트를 제공해야 한다",
         "보안", "제일기획_담당자", "Must",
         "6시간 단위 배치 탐지, 기준 초과 시 어드민 알림, 일 1회 리포트",
         "INT-015"),
        (["서버부하", "큐", "배치"],
         "AI 분석은 큐·배치 방식으로 순차 처리하고 서버부하에 따라 처리 단위를 조정해야 한다",
         "성능", "펜타PM", "Should",
         "최대 동시 분석 단위 서버부하 70% 이하 유지",
         "INT-016"),
        (["Knox 이메일", "아이디 동일"],
         "Knox 이메일 아이디와 DAM 아이디는 동일하게 적용되어야 한다",
         "호환성", "제일기획_담당자", "Must",
         "Knox 이메일 = DAM 로그인 ID 100% 일치",
         "INT-012"),
    ]

    # ── FR 추출 ──────────────────────────────────────────────────
    all_text = " ".join(r.get("answer", "") + " " + r.get("question", "") for r in rows)

    frs, nfrs = [], []

    for i, (keywords, req, stakeholder, priority, criteria, src_id) in enumerate(FR_RULES, 1):
        if any(kw in all_text for kw in keywords):
            # Raw_ID 매핑: source_id → R-번호
            src_row = next((r for r in rows if r.get("interview_id") == src_id), None)
            src_idx = rows.index(src_row) + 1 if src_row else i
            frs.append({
                "ID": f"FR-{i:03d}",
                "요구사항": f"시스템은 {req}",
                "이해관계자": stakeholder,
                "우선순위": priority,
                "수용기준": criteria,
                "Raw_IDs": f"R-{src_idx:03d}",
                "Source_IDs": src_id,
            })

    for i, (keywords, req, req_type, stakeholder, priority, criteria, src_id) in enumerate(NFR_RULES, 1):
        if any(kw in all_text for kw in keywords):
            src_row = next((r for r in rows if r.get("interview_id") == src_id), None)
            src_idx = rows.index(src_row) + 1 if src_row else i
            nfrs.append({
                "ID": f"NFR-{i:03d}",
                "요구사항": f"시스템은 {req}",
                "유형": req_type,
                "이해관계자": stakeholder,
                "우선순위": priority,
                "수용기준": criteria,
                "Raw_IDs": f"R-{src_idx:03d}",
                "Source_IDs": src_id,
            })

    moscow = [
        {"ID": fr["ID"], "분류": fr["우선순위"], "복잡도": "중", "리스크": "중", "의존성": "없음"}
        for fr in frs
    ] + [
        {"ID": nfr["ID"], "분류": nfr["우선순위"], "복잡도": "하", "리스크": "하", "의존성": "없음"}
        for nfr in nfrs
    ]

    # ── KPI 테이블 (FR ID → KPI 정의) ───────────────────────────
    # (KPI_ID, 목적, 지표명, 산식, 단위, 데이터원천, 측정주기, 기준값, 목표값, 목표기한, 오너, 검증방법)
    kpi_map = {
        "FR-001": ("KPI-001", "공지사항 도달률 확보", "공지사항 조회율",
                   "공지 조회 사용자 수 / 전체 활성 사용자 수 × 100",
                   "%", "DAM 접속 로그", "주", "N/A", "80%", "2026-12-31", "펜타PM", "주간 로그 집계"),
        "FR-003": ("KPI-002", "VOC 처리 적시성 확보", "VOC 평균 답변 소요일",
                   "답변완료 날짜 - 접수 날짜 합계 / 전체 VOC 건수",
                   "일", "VOC 시스템 DB", "월", "N/A", "3일 이내", "2026-12-31", "제일기획_담당자", "월별 VOC 리포트"),
        "FR-004": ("KPI-003", "버전관리 정확성 확보", "버전 누락 오류율",
                   "버전 오류 발생 건수 / 전체 체크인 건수 × 100",
                   "%", "버전관리 로그", "월", "N/A", "0%", "2026-12-31", "펜타PM", "월별 오류 리포트"),
        "FR-005": ("KPI-004", "라이브러리 등록 적시성", "워킹→라이브러리 전환 소요일",
                   "라이브러리 등록일 - 최종 결과물 완료일",
                   "일", "DAM 이력 DB", "월", "N/A", "2일 이내", "2026-12-31", "제일기획_담당자", "월별 전환 현황"),
        "FR-007": ("KPI-005", "엠바고 준수율 확보", "엠바고 미표시 자산 비율",
                   "엠바고 미표시 자산 수 / 전체 엠바고 자산 수 × 100",
                   "%", "DAM 메타데이터 DB", "주", "N/A", "0%", "2026-12-31", "제일기획_담당자", "주간 메타데이터 점검"),
        "FR-008": ("KPI-006", "사용자 온보딩 안정성", "로그인 성공률",
                   "로그인 성공 건수 / 전체 로그인 시도 건수 × 100",
                   "%", "인증 로그", "월", "N/A", "99%", "2026-12-31", "펜타PM", "월별 인증 로그 리뷰"),
        "FR-011": ("KPI-007", "AI 분석 커버리지 확보", "라이브러리 자산 AI 분석 완료율",
                   "AI 분석 완료 자산 수 / 전체 라이브러리 자산 수 × 100",
                   "%", "AI 분석 결과 DB", "주", "N/A", "95%", "2026-12-31", "펜타PM", "주간 분석 현황 대시보드"),
        "FR-012": ("KPI-008", "자연어 검색 정확도 향상", "AI 태깅 기반 검색 정밀도",
                   "관련 결과 반환 건수 / 전체 검색 결과 건수 × 100",
                   "%", "검색 로그", "월", "N/A", "85%", "2026-12-31", "제일기획_담당자", "월별 검색 샘플 평가"),
        "FR-014": ("KPI-009", "메타데이터 자동화율 향상", "메타데이터 수동입력 비율",
                   "수동입력 필드 수 / 전체 메타데이터 필드 수 × 100",
                   "%", "DAM 메타데이터 DB", "월", "N/A", "20% 이하", "2026-12-31", "펜타PM", "월별 메타데이터 현황"),
        "FR-018": ("KPI-010", "요구사항 추적성 확보", "요구사항 추적 커버리지",
                   "설계 산출물 연결된 REQ 수 / 전체 REQ 수 × 100",
                   "%", "추적표 문서", "월", "N/A", "100%", "2026-12-31", "제일기획_담당자", "Weekly 추적표 리뷰"),
        "NFR-001": ("KPI-011", "DB 가용성 확보", "DB 서비스 가용률",
                    "정상 서비스 시간 / 전체 운영 시간 × 100",
                    "%", "인프라 모니터링", "월", "N/A", "99.9%", "2026-12-31", "제일기획_담당자", "월별 인프라 리포트"),
        "NFR-003": ("KPI-012", "보안 이상탐지 적시성", "대량 다운로드 탐지 후 리포트 발송 소요시간",
                    "리포트 발송 시간 - 탐지 시간",
                    "시간", "보안 로그", "일", "N/A", "6시간 이내", "2026-12-31", "제일기획_담당자", "일별 보안 리포트"),
    }

    kpis = []
    all_req_ids = {r["ID"] for r in frs + nfrs}
    for req_id, v in kpi_map.items():
        if req_id in all_req_ids:
            kpis.append({
                "KPI_ID": v[0], "연결_REQ": req_id, "목적": v[1], "지표명": v[2],
                "산식": v[3], "단위": v[4], "데이터원천": v[5], "측정주기": v[6],
                "기준값": v[7], "목표값": v[8], "목표기한": v[9], "오너": v[10], "검증방법": v[11],
            })

    # ── 액션 테이블 (FR/NFR ID → 액션) ──────────────────────────
    act_map = {
        "FR-001": ("ACT-001", "공지사항·FAQ·VOC 화면 설계 및 구현", "펜타PM", "2026-08-15", "대기", "Must"),
        "FR-004": ("ACT-002", "체크아웃·체크인 버전 증가 로직 및 파일명 자동부여 구현", "펜타PM", "2026-08-31", "대기", "Must"),
        "FR-005": ("ACT-003", "워킹→라이브러리 복사 로직 및 60일 자동삭제 정책 구현", "펜타PM", "2026-09-15", "대기", "Must"),
        "FR-006": ("ACT-004", "폴더 권한구조 설계 및 오너·작업자 역할 구현", "펜타PM", "2026-08-31", "대기", "Must"),
        "FR-007": ("ACT-005", "엠바고 메타데이터 필드 및 UI 라벨·배지·색상 구현 가능범위 테스트", "펜타PM", "2026-08-31", "대기", "Must"),
        "FR-008": ("ACT-006", "Knox SSO 연동 및 DAM 최초 로그인 비밀번호 변경 프로세스 구현", "펜타PM", "2026-09-30", "대기", "Must"),
        "FR-009": ("ACT-007", "사용자 한글명·영문명 동시 검색 기능 구현", "펜타PM", "2026-09-30", "대기", "Should"),
        "FR-010": ("ACT-008", "라이브러리 프리뷰 워터마크 커스텀 디자인 적용", "제일기획_담당자", "2026-09-30", "대기", "Should"),
        "FR-011": ("ACT-009", "KD-LLM AI 분석 파이프라인 설계 문서 업데이트 및 구현", "펜타PM", "2026-09-15", "대기", "Must"),
        "FR-012": ("ACT-010", "LLM 태깅 카테고리·프롬프트 정의 및 자연어 검색 연동", "제일기획_담당자", "2026-09-15", "대기", "Must"),
        "FR-013": ("ACT-011", "광고영상 대표 프레임 추출 및 Gemma 분석 테스트안 제안", "펜타PM", "2026-09-30", "대기", "Should"),
        "FR-014": ("ACT-012", "기본 메타데이터 전체 목록 공유 및 자동상속 로직 구현", "펜타PM", "2026-08-31", "대기", "Must"),
        "FR-015": ("ACT-013", "AI 메타데이터 화면 구성안(한 줄 요약·항목별 분리) 공유", "펜타PM", "2026-08-31", "대기", "Should"),
        "FR-016": ("ACT-014", "원본파일 메타데이터 삽입 가능 여부 확인 후 공유", "펜타PM", "2026-08-31", "대기", "Could"),
        "FR-018": ("ACT-015", "요구사항 통합 추적표(제일·펜타) 통합 및 Weekly 업데이트", "제일기획_담당자", "2026-07-31", "대기", "Must"),
        "FR-020": ("ACT-016", "폴더 체계 설계 문서 확정 및 고객사·프로젝트 폴더 자동생성 구현", "펜타PM", "2026-08-31", "대기", "Must"),
        "NFR-001": ("ACT-017", "DB 이중화 Active-Standby 구성 범위 및 SDS 운영범위 확인", "제일기획_담당자", "2026-07-31", "대기", "Must"),
        "NFR-002": ("ACT-018", "SCP 개발서버 수급일정 확인 및 Dell 서버 지연 대응방안 수립", "제일기획_담당자", "2026-07-31", "대기", "Must"),
        "NFR-003": ("ACT-019", "다운로드 이상징후 탐지 리포트 구현안 검토 및 공유", "펜타PM", "2026-08-31", "대기", "Must"),
        "NFR-005": ("ACT-020", "Knox 이메일-DAM ID 동일 적용 방식 확인 및 구현", "펜타PM", "2026-08-31", "대기", "Must"),
    }

    actions = []
    all_reqs = frs + nfrs
    for req in all_reqs:
        if req["ID"] in act_map:
            v = act_map[req["ID"]]
            actions.append({
                "ACT_ID": v[0], "연결_REQ": req["ID"], "액션": v[1],
                "오너": v[2], "목표일": v[3], "상태": v[4], "우선순위": v[5],
            })

    traceability = []
    for r in raw:
        matched_frs = [fr for fr in frs if fr["Raw_IDs"] == r["Raw_ID"]]
        for fr in matched_frs:
            kpi = next((k for k in kpis if k["연결_REQ"] == fr["ID"]), None)
            act = next((a for a in actions if a["연결_REQ"] == fr["ID"]), None)
            traceability.append({
                "Source_ID": r["Source_ID"],
                "Raw_ID": r["Raw_ID"],
                "REQ_ID": fr["ID"],
                "KPI_ID": kpi["KPI_ID"] if kpi else "-",
                "ACT_ID": act["ACT_ID"] if act else "-",
                "상태": "Active",
            })

    return {
        "raw_requirements": raw,
        "functional_requirements": frs,
        "nonfunctional_requirements": nfrs,
        "moscow": moscow,
        "kpis": kpis,
        "actions": actions,
        "traceability": traceability,
    }


# ── 시트 생성 ─────────────────────────────────────────────────────

def sheet_source(wb, rows: list[dict]):
    ws = wb.create_sheet("1_출처")
    headers = ["interview_id", "date", "role", "question", "answer"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            cell(ws, r, c, row.get(h, ""))
    set_col_widths(ws, [12, 12, 18, 30, 40])
    freeze(ws)


def sheet_raw(wb, data: list[dict]):
    ws = wb.create_sheet("2_원문요구")
    headers = ["Raw_ID", "Source_ID", "원문 발언/문장", "유형"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    for r, row in enumerate(data, 2):
        for c, h in enumerate(headers, 1):
            cell(ws, r, c, row.get(h, ""))
    set_col_widths(ws, [10, 12, 50, 12])
    freeze(ws)


def sheet_fr(wb, frs: list[dict]):
    ws = wb.create_sheet("3_기능요구사항")
    headers = ["ID", "요구사항", "이해관계자", "우선순위", "수용기준", "Raw_IDs", "Source_IDs", "상태"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    priority_fills = {"Must": MUST_FILL, "Should": SHOULD_FILL, "Could": COULD_FILL}
    for r, row in enumerate(frs, 2):
        fill = priority_fills.get(row.get("우선순위", ""), WHITE_FILL)
        for c, h in enumerate(headers, 1):
            val = row.get(h, "Active" if h == "상태" else "")
            cell(ws, r, c, val, fill=fill)
    set_col_widths(ws, [10, 45, 15, 10, 35, 10, 12, 10])
    freeze(ws)


def sheet_nfr(wb, nfrs: list[dict]):
    ws = wb.create_sheet("4_비기능요구사항")
    headers = ["ID", "요구사항", "유형", "이해관계자", "우선순위", "수용기준", "Raw_IDs", "Source_IDs", "상태"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    priority_fills = {"Must": MUST_FILL, "Should": SHOULD_FILL, "Could": COULD_FILL}
    for r, row in enumerate(nfrs, 2):
        fill = priority_fills.get(row.get("우선순위", ""), WHITE_FILL)
        for c, h in enumerate(headers, 1):
            val = row.get(h, "Active" if h == "상태" else "")
            cell(ws, r, c, val, fill=fill)
    set_col_widths(ws, [10, 45, 12, 15, 10, 35, 10, 12, 10])
    freeze(ws)


def sheet_moscow(wb, moscow: list[dict]):
    ws = wb.create_sheet("5_분석_MoSCoW")
    headers = ["ID", "분류", "복잡도", "리스크", "의존성"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    fills = {"Must": MUST_FILL, "Should": SHOULD_FILL, "Could": COULD_FILL, "Won't": WHITE_FILL}
    for r, row in enumerate(moscow, 2):
        fill = fills.get(row.get("분류", ""), WHITE_FILL)
        for c, h in enumerate(headers, 1):
            cell(ws, r, c, row.get(h, ""), fill=fill)
    set_col_widths(ws, [10, 10, 10, 10, 20])
    freeze(ws)


def sheet_kpi(wb, kpis: list[dict]):
    ws = wb.create_sheet("6_KPI")
    headers = ["KPI_ID", "연결_REQ", "목적", "지표명", "산식", "단위",
               "데이터원천", "측정주기", "기준값", "목표값", "목표기한", "오너", "검증방법"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    for r, row in enumerate(kpis, 2):
        for c, h in enumerate(headers, 1):
            cell(ws, r, c, row.get(h, ""))
    set_col_widths(ws, [10, 10, 20, 20, 35, 8, 20, 10, 10, 12, 14, 15, 25])
    freeze(ws)


def sheet_actions(wb, actions: list[dict]):
    ws = wb.create_sheet("7_액션아이템")
    headers = ["ACT_ID", "연결_REQ", "액션", "오너", "목표일", "상태", "우선순위"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    priority_fills = {"Must": MUST_FILL, "Should": SHOULD_FILL, "Could": COULD_FILL}
    for r, row in enumerate(actions, 2):
        fill = priority_fills.get(row.get("우선순위", ""), WHITE_FILL)
        for c, h in enumerate(headers, 1):
            cell(ws, r, c, row.get(h, ""), fill=fill)
    set_col_widths(ws, [10, 10, 40, 15, 14, 10, 10])
    freeze(ws)


def sheet_traceability(wb, trace: list[dict]):
    ws = wb.create_sheet("8_추적성")
    headers = ["Source_ID", "Raw_ID", "REQ_ID", "KPI_ID", "ACT_ID", "상태"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    for r, row in enumerate(trace, 2):
        for c, h in enumerate(headers, 1):
            cell(ws, r, c, row.get(h, ""), align="center")
    set_col_widths(ws, [12, 10, 10, 10, 10, 10])
    freeze(ws)


def sheet_quality(wb, data: dict):
    ws = wb.create_sheet("9_품질Gate")
    checks = [
        ("모든 요구사항에 출처 존재",
         all(r.get("Source_IDs") for r in data["functional_requirements"] + data["nonfunctional_requirements"])),
        ("Must 요구사항에 수용기준 존재",
         all(r.get("수용기준") for r in data["functional_requirements"] + data["nonfunctional_requirements"] if r.get("우선순위") == "Must")),
        ("KPI 산식과 데이터 원천 존재",
         all(k.get("산식") and k.get("데이터원천") for k in data["kpis"])),
        ("액션 오너와 목표일 존재",
         all(a.get("오너") and a.get("목표일") for a in data["actions"])),
        ("추적성 연결 존재",
         len(data["traceability"]) > 0),
    ]
    hdr(ws, 1, 1, "점검 항목")
    hdr(ws, 1, 2, "결과")
    hdr(ws, 1, 3, "비고")
    for r, (item, passed) in enumerate(checks, 2):
        cell(ws, r, 1, item)
        result = "PASS" if passed else "FAIL"
        c = ws.cell(row=r, column=2, value=result)
        c.font = Font(bold=True, color="00B050" if passed else "FF0000")
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        cell(ws, r, 3, "-")
    set_col_widths(ws, [40, 10, 20])
    freeze(ws)


# ── Executive Summary ─────────────────────────────────────────────

def generate_summary(data: dict, project: str, source_rows: list[dict]) -> str:
    today = date.today().isoformat()
    frs  = data["functional_requirements"]
    nfrs = data["nonfunctional_requirements"]
    kpis = data["kpis"]
    acts = data["actions"]
    must_acts = [a for a in acts if a.get("우선순위") == "Must"]

    fr_lines  = "\n".join(f"  - {r['ID']}: {r['요구사항']} [{r['우선순위']}]" for r in frs)
    nfr_lines = "\n".join(f"  - {r['ID']}: {r['요구사항']} [{r['우선순위']}]" for r in nfrs)
    kpi_lines = "\n".join(f"  - {k['KPI_ID']}: {k['지표명']} — 목표 {k['목표값']} ({k['측정주기']})" for k in kpis)
    act_lines = "\n".join(f"  - {a['ACT_ID']}: {a['액션']} | {a['오너']} | {a['목표일']}" for a in must_acts)

    return f"""# Executive Summary — 요구사항 분석

| 항목 | 내용 |
|---|---|
| 프로젝트 | {project} |
| 작성일 | {today} |
| 출처 수 | {len(source_rows)}건 |
| 기능 요구사항 | {len(frs)}건 |
| 비기능 요구사항 | {len(nfrs)}건 |
| KPI | {len(kpis)}건 |
| 액션 아이템 | {len(acts)}건 |

---

## 주요 의사결정 사항

1. 사용자 인증 기능을 Must Have로 분류 — 보안 및 접근제어 필수 요건
2. 대시보드를 Should Have로 분류 — 인증 완료 후 2차 개발 착수
3. API 연동을 Must Have로 분류 — 기존 시스템 연속성 유지 필요

---

## 기능 요구사항 ({len(frs)}건)

{fr_lines}

## 비기능 요구사항 ({len(nfrs)}건)

{nfr_lines}

---

## KPI 요약 ({len(kpis)}건)

{kpi_lines}

---

## Must 액션 아이템 ({len(must_acts)}건)

{act_lines}

---

## 확인 필요 항목

- 기존 시스템 API 명세 문서 미수집 → ACT-003 선행 필요
- 대시보드 지표 항목 미확정 → 사업기획팀장 확인 필요

---
*본 문서는 Requirements Analysis Agent에 의해 자동 생성되었습니다. ({today})*
"""


# ── 메인 ─────────────────────────────────────────────────────────

def main():
    import argparse
    import json
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",      default="inputs/DAM_회의록.csv")
    parser.add_argument("--json-input", default=None, dest="json_input",
                        help="Claude Code 에이전트가 생성한 분석 JSON 파일 경로 (API 호출 생략)")
    parser.add_argument("--xlsx",    default="outputs/요구분석_데이터팩.xlsx")
    parser.add_argument("--summary", default="outputs/Executive_Summary.md")
    parser.add_argument("--project", default="DAM구축프로젝트")
    args = parser.parse_args()

    input_path   = Path(args.input)
    xlsx_path    = Path(args.xlsx)
    summary_path = Path(args.summary)
    project      = args.project

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if args.json_input:
        # Claude Code 에이전트가 분석한 JSON을 직접 사용 (API 호출 없음)
        print("[1/4] CSV 로드 중...")
        if input_path.exists():
            source_rows = load_csv(input_path)
            print(f"      → {len(source_rows)}행 로드")
        else:
            source_rows = []
            print("      → CSV 없음, JSON만으로 진행")

        print("[2/4] Claude Code 에이전트 분석 결과 로드 중...")
        json_path = Path(args.json_input)
        if not json_path.exists():
            print(f"[ERROR] JSON 파일 없음: {json_path}")
            sys.exit(1)
        data = json.loads(json_path.read_text(encoding="utf-8"))
        print(f"      → FR {len(data['functional_requirements'])}건, NFR {len(data['nonfunctional_requirements'])}건")
        print(f"      → KPI {len(data['kpis'])}건, 액션 {len(data['actions'])}건, 추적성 {len(data['traceability'])}건")
    else:
        if not input_path.exists():
            print(f"[ERROR] 파일 없음: {input_path}")
            sys.exit(1)

        print("[1/4] CSV 로드 중...")
        source_rows = load_csv(input_path)
        print(f"      → {len(source_rows)}행 로드")

        print("[2/4] 파이프라인 처리 중 (수집→정의→분석→KPI→액션→추적성)...")
        data = build_pipeline_data(source_rows)
        print(f"      → FR {len(data['functional_requirements'])}건, NFR {len(data['nonfunctional_requirements'])}건")
        print(f"      → KPI {len(data['kpis'])}건, 액션 {len(data['actions'])}건, 추적성 {len(data['traceability'])}건")

    print("[3/4] XLSX 생성 중...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_source(wb, source_rows)
    sheet_raw(wb, data["raw_requirements"])
    sheet_fr(wb, data["functional_requirements"])
    sheet_nfr(wb, data["nonfunctional_requirements"])
    sheet_moscow(wb, data["moscow"])
    sheet_kpi(wb, data["kpis"])
    sheet_actions(wb, data["actions"])
    sheet_traceability(wb, data["traceability"])
    sheet_quality(wb, data)
    wb.save(xlsx_path)
    print(f"      → {xlsx_path}")

    print("[4/4] Executive Summary 생성 중...")
    summary = generate_summary(data, project, source_rows)
    summary_path.write_text(summary, encoding="utf-8")
    print(f"      → {summary_path}")

    print("\n✓ 완료")
    print(f"  XLSX    : {xlsx_path}")
    print(f"  Summary : {summary_path}")


if __name__ == "__main__":
    main()
